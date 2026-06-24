#!/usr/bin/env python3
"""Xac minh chinh xac so dong data - dem tung dong"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl

docs_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'docs')

def analyze_file(fpath, fname):
    print(f'\n{"="*70}')
    print(f'FILE: {fname}')
    print(f'{"="*70}')

    wb = openpyxl.load_workbook(fpath, data_only=True)

    for sname in wb.sheetnames:
        safe = sname.encode('ascii', errors='replace').decode()
        ws = wb[sname]

        # Dem TUNG DONG thuc su
        total = 0
        empty = 0
        header_row = None
        sample_rows = []

        for row in ws.iter_rows(min_row=1, values_only=True):
            total += 1
            vals = [str(v or '').strip() for v in row]
            if not any(v for v in vals):
                empty += 1
                continue

            # Tim header: dong dau tien co >= 2 cot chua text
            if header_row is None:
                text_cols = sum(1 for v in vals if v and any(c.isalpha() for c in v))
                if text_cols >= 2:
                    header_row = total
                    continue  # skip header from data

            # Data row
            if total > (header_row or 0):
                sample_rows.append(vals)

        data_count = total - empty - (1 if header_row else 0)
        print(f'  [{safe}] total_rows={total} | empty={empty} | header_row={header_row} | data_rows={data_count}')

        if sample_rows:
            # Show 1st and last 2 data rows
            for label, idx in [('First', 0), ('Last-1', max(0, len(sample_rows)-2)), ('Last', -1)]:
                if idx < len(sample_rows) and idx >= 0:
                    row = sample_rows[idx]
                    non_empty = [(chr(65+i) if i<26 else str(i), v[:60]) for i, v in enumerate(row) if v]
                    print(f'    {label} data (row {header_row+1+idx}): {non_empty[:8]}')
    wb.close()

# Phan tich tung file
for fname in sorted(os.listdir(docs_dir)):
    if fname.endswith('.xlsx') and not fname.startswith('~$'):
        analyze_file(os.path.join(docs_dir, fname), fname)
