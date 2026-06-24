#!/usr/bin/env python3
"""Phan tich toan bo file Excel trong docs/"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl

docs_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'docs')
files = [f for f in os.listdir(docs_dir) if f.endswith('.xlsx') and not f.startswith('~$')]

for fname in sorted(files):
    fpath = os.path.join(docs_dir, fname)
    size_kb = os.path.getsize(fpath) / 1024
    print('=' * 70)
    print(f'FILE: {fname}')
    print(f'  Dung luong: {size_kb:,.0f} KB')

    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    safe_names = [s.encode('ascii', errors='replace').decode() for s in wb.sheetnames]
    print(f'  So sheet: {len(wb.sheetnames)}')
    print(f'  Sheets: {safe_names}')
    print()

    for sname in wb.sheetnames:
        safe_sname = sname.encode('ascii', errors='replace').decode()
        ws = wb[sname]
        total_rows = ws.max_row
        total_cols = ws.max_column

        # Read first 8 rows
        rows_data = []
        for row in ws.iter_rows(min_row=1, max_row=min(8, total_rows or 1), values_only=True):
            rows_data.append([str(v or '')[:80] for v in row])

        print(f'  --- Sheet: "{safe_sname}" ---')
        print(f'  Dong: {total_rows} | Cot: {total_cols}')

        if not rows_data:
            print('  [EMPTY SHEET]\n')
            continue

        # Determine best header row among first 4 rows
        header_row_idx = 0
        best_count = 0
        for ri in range(min(4, len(rows_data))):
            cnt = sum(1 for v in rows_data[ri] if v.strip())
            if cnt > best_count:
                best_count = cnt
                header_row_idx = ri

        print(f'  HEADER (row {header_row_idx+1}, {best_count}/{total_cols} cols filled):')
        for i, v in enumerate(rows_data[header_row_idx]):
            if v.strip():
                col_letter = chr(65+i) if i < 26 else f'Col{i}'
                safe_v = v.encode('ascii', errors='replace').decode()
                print(f'    {col_letter}: "{safe_v}"')

        # Sample data
        data_rows = [r for ri, r in enumerate(rows_data) if ri > header_row_idx]
        if data_rows:
            print(f'  SAMPLE DATA:')
            sample = data_rows[0]
            for i, v in enumerate(sample[:len(rows_data[header_row_idx])]):
                if v.strip():
                    col_letter = chr(65+i) if i < 26 else f'Col{i}'
                    safe_v = v.encode('ascii', errors='replace').decode()
                    print(f'    {col_letter}: "{safe_v}"')

        if total_rows and total_rows > header_row_idx + 1:
            est = total_rows - header_row_idx - 1
            print(f'  -> ~{est} data rows')
        print()
    wb.close()
    print()
