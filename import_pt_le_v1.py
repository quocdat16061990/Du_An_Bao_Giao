"""
Import products from BAO_GIA_PT_LE 26-07.xlsx with CORRECT column mapping.
- Backs up existing images (hinh_anh, danh_sach_hinh_anh) by ma_vt
- Deletes existing products for matched types
- Re-imports from Excel
- Restores images to newly created products

Usage: py import_pt_le_v1.py --dry-run | py import_pt_le_v1.py
"""

import sys, os, re, argparse, unicodedata
from pathlib import Path
from decimal import Decimal, InvalidOperation

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'backend.settings')
env_file = Path(__file__).resolve().parent / 'backend' / '.env'
if env_file.exists():
    for line in env_file.read_text(encoding='utf-8').splitlines():
        line = line.strip()
        if line and not line.startswith('#') and '=' in line:
            key, _, val = line.partition('=')
            os.environ.setdefault(key.strip(), val.strip())

import django; django.setup()
from products.models import Product, HangMay, HangSx, ThuongHieu, Category
from django.utils.text import slugify
import openpyxl


DOCS_DIR = Path(__file__).resolve().parent / 'docs' / 'update-docs-v1'
EXCEL_FILE = DOCS_DIR / 'BAO_GIA_PT_LE 26-07.xlsx'

# Sheet name -> loai
SHEET_MAP = {
    'SUPAP': 'supap',
    'TRUC CO': 'truc_co',
    'BOM NUOC': 'bom_nuoc',
    'NAP QUY LAT': 'nap_quy_lat',
    'BOM NHOT': 'bom_nhot',
    'TRUC CAM': 'truc_cam',
    'NAP SINH HAN': 'nap_sinh_han',
    'RUOT SINH HAN': 'ruot_sinh_han',
    'TAY BIEN': 'nhip_tay_bien',
    'THUN RON': 'thun_co',
    'THUN XY LANH': 'thun_xy_lanh',
    'LOC MAY': 'loc_may',
    'SAM BEC': 'sam_bac',
    'VAN HANG NHIET': 'van_hang_nhiet',
    'VANH RANG BANH DA': 'vanh_rang_banh_da',
    'ONG DAN NHIEN LIEU': 'ong_dan_nhien_lieu',
    'SEN CAM CO BO LUA': 'sen_cam',
    'NAP NUOC NAP CAM SAO': 'nap_quy_lat',
    'MAY PHAT DINAMO': 'bom_nuoc',
}


def parse_price(val) -> Decimal | None:
    if val is None: return None
    if isinstance(val, (int, float)):
        return Decimal(str(int(val)))
    text = str(val).strip()
    if not text or text.lower() in ('none', 'Ã¢â‚¬â€', '-', ''): return None
    clean = re.sub(r'[Ã¢â€šÂ«Ã„â€˜VND\s,]', '', text)
    if '.' in clean and ',' in clean:
        clean = clean.replace('.', '').replace(',', '.')
    elif '.' in clean:
        parts = clean.split('.')
        if len(parts) > 1 and len(parts[-1]) == 3: clean = clean.replace('.', '')
    try:
        return Decimal(clean)
    except InvalidOperation:
        return None

def get_or_create(model_class, ten: str):
    if not ten or ten in ('None', 'Ã¢â‚¬â€', '', ' '): return None
    obj = model_class.objects.filter(ten__iexact=ten).first()
    if obj: return obj
    base_slug = slugify(ten)
    slug = base_slug
    counter = 1
    while True:
        try:
            return model_class.objects.create(ten=ten, slug=slug)
        except Exception:
            existing = model_class.objects.filter(slug=slug).first()
            if existing: return existing
            counter += 1
            slug = f'{base_slug}-{counter}'
            if counter > 100: return None

def is_numeric(s):
    return bool(re.match(r'^[\d.]+$', s.strip()))


def normalize_header(value):
    text = str(value or '').replace(chr(272), 'D').replace(chr(273), 'd')
    text = unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('ascii')
    return re.sub(r'[^A-Z0-9]+', ' ', text.upper()).strip()


def get_sheet_columns(ws):
    columns = {}
    for column in range(1, ws.max_column + 1):
        header = normalize_header(ws.cell(2, column).value)
        header_map = {
            'MA HH': 'ma_hh', 'HANG MAY': 'hang_may', 'TEN DONG CO': 'ten_sp',
            'DUONG KINH PISTON': 'duong_kinh_piston', 'NAM X THAN X DAI': 'kich_thuoc',
            'PARNO': 'parno', 'GIA VIP': 'gia_vip', 'GIA UU DAI': 'gia_uu_dai',
            'GIA DAI LY': 'gia_dai_ly', 'GIA GARA': 'gia_gara',
            'DON VI': 'dvt', 'GHI CHU': 'ghi_chu',
        }
        for marker, key in header_map.items():
            if marker in header:
                columns[key] = column
                break
    return columns


def cell_value(ws, row, columns, key):
    column = columns.get(key)
    return ws.cell(row, column).value if column else None


def cell_text(ws, row, columns, key):
    return str(cell_value(ws, row, columns, key) or '').strip()

def import_pt_le(dry_run=True):
    if not EXCEL_FILE.exists():
        print(f'[ERROR] File not found: {EXCEL_FILE}')
        return

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    types_to_process = list(set(SHEET_MAP.values()))

    # Phase 1: Backup images
    print("--- Backing up images ---")
    image_map = {}
    products_to_delete = Product.objects.filter(loai__in=types_to_process)
    for p in products_to_delete:
        if p.hinh_anh or p.danh_sach_hinh_anh:
            image_map[p.ma_vt] = {
                'hinh_anh': p.hinh_anh,
                'danh_sach_hinh_anh': p.danh_sach_hinh_anh
            }
    print(f"Backed up images for {len(image_map)} products.")

    # Phase 2: Delete existing
    print(f'Products to DELETE ({len(types_to_process)} types): {products_to_delete.count()}')
    if not dry_run:
        deleted, _ = products_to_delete.delete()
        print(f'Deleted: {deleted}')

    # Phase 3: Import
    stats = {'created': 0, 'restored_images': 0, 'skipped': 0, 'errors': 0}

    for sname in wb.sheetnames:
        sheet_key = normalize_header(sname)
        loai = SHEET_MAP.get(sheet_key)
        if not loai:
            print(f'[WARN] Sheet not mapped: {sname}')
            continue

        ws = wb[sname]
        print(f'\n--- Sheet: {sname.strip()} -> loai: {loai} ---')
        columns = get_sheet_columns(ws)
        required_columns = {'ma_hh', 'hang_may', 'ten_sp'}
        if not required_columns.issubset(columns):
            print(f'[WARN] Invalid columns in sheet {sname}: {columns}')
            continue

        for row in range(4, ws.max_row + 1):
            ma_hh = cell_text(ws, row, columns, 'ma_hh')
            if not ma_hh.startswith('HH'):
                continue

            hang_may_raw = cell_text(ws, row, columns, 'hang_may')
            ten_sp = cell_text(ws, row, columns, 'ten_sp')
            parno = cell_text(ws, row, columns, 'parno')
            duong_kinh_piston = cell_text(ws, row, columns, 'duong_kinh_piston')
            kich_thuoc = cell_text(ws, row, columns, 'kich_thuoc')
            dvt = cell_text(ws, row, columns, 'dvt')
            ghi_chu = cell_text(ws, row, columns, 'ghi_chu')
            gia_vip = parse_price(cell_value(ws, row, columns, 'gia_vip'))
            gia_uu_dai = parse_price(cell_value(ws, row, columns, 'gia_uu_dai'))
            gia_dai_ly = parse_price(cell_value(ws, row, columns, 'gia_dai_ly'))
            gia_gara = parse_price(cell_value(ws, row, columns, 'gia_gara'))
            attributes = {}
            if duong_kinh_piston:
                attributes['duong_kinh_piston'] = duong_kinh_piston
            if kich_thuoc:
                attributes['kich_thuoc'] = kich_thuoc

            if not hang_may_raw or hang_may_raw in ('None', '???', ''):
                stats['skipped'] += 1
                continue

            hang_may_raw = hang_may_raw.replace('Ã¢â€â€Ã¢â€â‚¬', '').strip()

            if dry_run:
                stats['created'] += 1
                continue

            hm = get_or_create(HangMay, ten=hang_may_raw)
            if not hm:
                stats['skipped'] += 1
                continue

            try:
                obj = Product.objects.create(
                    ma_vt=ma_hh,
                    loai=loai,
                    ten_hang=ten_sp[:500] if ten_sp else '',
                    hang_may=hm,
                    parno=parno[:300] if parno else '',
                    dvt=dvt[:50] if dvt else 'C?i',
                    gia_vip=gia_vip,
                    gia_uu_dai=gia_uu_dai,
                    gia_dai_ly=gia_dai_ly,
                    gia_gara=gia_gara,
                    ghi_chu=ghi_chu or '',
                    attributes=attributes,
                    sheet_name=sname[:50],
                    is_active=True,
                )
                stats['created'] += 1

                # Restore images
                if ma_hh in image_map:
                    obj.hinh_anh = image_map[ma_hh]['hinh_anh']
                    obj.danh_sach_hinh_anh = image_map[ma_hh]['danh_sach_hinh_anh']
                    obj.save(update_fields=['hinh_anh', 'danh_sach_hinh_anh'])
                    stats['restored_images'] += 1
            except Exception as e:
                stats['errors'] += 1
                if stats['errors'] <= 3:
                    print(f'  ERROR [{ma_hh}]: {e}')

    wb.close()
    print(f"\nImport finished. Stats: {stats}")

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--dry-run', action='store_true', help='Run without modifying DB')
    args = parser.parse_args()

    import_pt_le(dry_run=args.dry_run)
