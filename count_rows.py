#!/usr/bin/env python3
"""Dem chinh xac so dong data tung sheet + kiem tra overlap"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl

docs_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'docs')
files = [f for f in os.listdir(docs_dir) if f.endswith('.xlsx') and not f.startswith('~$')]

for fname in sorted(files):
    fpath = os.path.join(docs_dir, fname)
    size_kb = os.path.getsize(fpath) / 1024
    print(f'\n{"="*70}')
    print(f'FILE: {fname} ({size_kb:.0f} KB)')
    print(f'{"="*70}')

    wb = openpyxl.load_workbook(fpath, data_only=True)

    for sname in wb.sheetnames:
        safe = sname.encode('ascii', errors='replace').decode()
        ws = wb[sname]

        # Count all rows with any data
        total_rows = 0
        data_rows = 0
        empty_rows = 0
        header_row = None
        header_cols = []

        # Read all rows to count
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
            vals = [str(v or '').strip() for v in row]
            has_data = any(v for v in vals)
            total_rows += 1

            # Detect header by finding row with most text
            if has_data:
                text_count = sum(1 for v in vals if v and any(c.isalpha() for c in v))
                if text_count >= 3:
                    if header_row is None:
                        header_row = row_idx
                        header_cols = vals

        # Count actual data rows (non-empty after header)
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
            vals = [str(v or '').strip() for v in row]
            has_data = any(v for v in vals)
            if has_data and (header_row is None or row_idx > header_row):
                # Check if it's a category separator (like "🏭 CAT · 34 loai")
                first_val = vals[0] if vals else ''
                if any(emoji in first_val for emoji in ['🏭','🚗','🔧','📦','⚙️','🛠️','🔩','📒','🚌','🏍️','🟢']):
                    continue
                # Check pattern "X · Y sản phẩm" or "X (Y products)"
                if ' sản phẩm)' in first_val or ' loại' in first_val:
                    continue
                data_rows += 1

        if header_cols:
            # Clean header
            clean_hdrs = [h[:40] for h in header_cols[:10] if h]
            print(f'  [{safe}] total_rows={total_rows} | data_rows={data_rows} | header={header_row}')
            print(f'    Header: {clean_hdrs}')
        else:
            print(f'  [{safe}] total_rows={total_rows} | data_rows={data_rows} | NO HEADER')
    wb.close()
