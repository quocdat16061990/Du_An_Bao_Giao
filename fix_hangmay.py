"""
Fix HangMay bị sai (số bore size thay vì tên hãng) bằng cách:
1. Đọc lại TONG_HOP_PHU_TUNG Excel — map Mã HH → Hãng Máy chuẩn
2. Đọc BAO_GIA_TURBO Excel — map Mã HH → Hãng Máy chuẩn (từ các sheet turbo)
3. Cập nhật HangMay cho sản phẩm trong DB
4. Các sản phẩm không fix được → gán "CHƯA RÕ"

Usage:
    py fix_hangmay.py --dry-run    # Preview
    py fix_hangmay.py              # Run for real
"""

import sys, os, re, argparse
from pathlib import Path
from collections import defaultdict

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'backend.settings')
env_file = Path(__file__).resolve().parent / 'backend' / '.env'
if env_file.exists():
    for line in env_file.read_text(encoding='utf-8').splitlines():
        line = line.strip()
        if line and not line.startswith('#') and '=' in line:
            key, _, val = line.partition('=')
            os.environ.setdefault(key.strip(), val.strip())

import django; django.setup()
from products.models import Product, HangMay
from django.utils.text import slugify
import openpyxl


DOCS_DIR = Path(__file__).resolve().parent / 'docs'


def is_numeric(s: str) -> bool:
    """Check if string is purely numeric (or decimal like 95.4, 117.9)."""
    return bool(re.match(r'^[\d.]+$', s.strip()))


def extract_brand_from_name(name: str) -> str | None:
    """Try to guess machine brand from product name."""
    if not name:
        return None
    n = name.upper()
    # Known brand patterns
    brands = [
        'KOMATSU', 'HITACHI', 'KOBELCO', 'CATERPILLAR', 'CAT',
        'HYUNDAI', 'VOLVO', 'DOOSAN', 'DAEWOO', 'SUMITOMO',
        'HINO', 'ISUZU', 'MITSUBISHI', 'NISSAN', 'TOYOTA',
        'CUMMINS', 'DEUTZ', 'PERKINS', 'PERKIN', 'YANMAR',
        'KUBOTA', 'DONGFENG', 'WEICHAI', 'YUCHAI', 'FAW',
        'SCANIA', 'MAN', 'MERCEDES', 'SUZUKI', 'MAZDA',
        'FORD', 'DETROIT', 'CHANGCHAI', 'QUANCHAI',
        'SHANG CHAI', 'SINOTRUCK', 'KIA', 'IHI',
    ]
    for b in sorted(brands, key=len, reverse=True):
        if b in n:
            return b
    return None


def parse_tong_hop_excel() -> dict[str, str]:
    """Parse TONG_HOP Excel: ma_hh -> hang_may_name"""
    mapping = {}
    fpath = DOCS_DIR / 'TONG_HOP_PHU_TUNG_10062026_2_updated.xlsx'
    if not fpath.exists():
        print(f'[WARN] File not found: {fpath}')
        return mapping

    wb = openpyxl.load_workbook(fpath, data_only=True)
    for sname in wb.sheetnames:
        ws = wb[sname]
        for row in range(4, ws.max_row + 1):
            ma_hh = str(ws.cell(row, 1).value or '').strip()
            hang_may = str(ws.cell(row, 2).value or '').strip()

            if not ma_hh or not hang_may:
                continue
            if ma_hh.startswith('◆') or ma_hh.startswith('▸'):
                continue

            # Only store if hang_may is a real brand name (not numeric, not empty)
            if hang_may and not is_numeric(hang_may) and hang_may not in ('None', '—', '-'):
                mapping[ma_hh] = hang_may

    wb.close()
    return mapping


def parse_turbo_excel() -> dict[str, str]:
    """Parse BAO_GIA_TURBO Excel: ma_vt -> hang_may_name"""
    mapping = {}
    fpath = DOCS_DIR / 'BAO_GIA_TURBO_v13_GIA_CO_ANH.xlsx'
    if not fpath.exists():
        print(f'[WARN] File not found: {fpath}')
        return mapping

    wb = openpyxl.load_workbook(fpath, data_only=True)
    for sname in ['🚗 BÁO GIÁ TURBO ', 'RUỘT TURBO', '🔩 SÒ & LINH KIỆN']:
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        for row in range(5, ws.max_row + 1):
            hang_may = str(ws.cell(row, 1).value or '').strip()
            ma_hh = str(ws.cell(row, 2).value or '').strip()

            if not ma_hh or not hang_may:
                continue
            # Skip category separator rows
            if '◆' in hang_may or '🏭' in hang_may or '─' in hang_may:
                continue
            if hang_may.startswith('└'):
                hang_may = hang_may[1:].strip()

            if hang_may and not is_numeric(hang_may) and hang_may not in ('None', '—', '-'):
                mapping[ma_hh] = hang_may

    wb.close()
    return mapping


def get_or_create_hangmay(name: str) -> HangMay:
    """Get or create HangMay by name."""
    hm = HangMay.objects.filter(ten__iexact=name).first()
    if hm:
        return hm
    hm = HangMay.objects.filter(ten__icontains=name).first()
    if hm:
        return hm
    hm = HangMay.objects.create(ten=name, slug=slugify(name))
    return hm


def fix_hangmay(dry_run: bool = True):
    """Main fix function."""
    # Build mapping from Excel files
    print('Building mapping from Excel files...')
    turbo_map = parse_turbo_excel()
    tonghop_map = parse_tong_hop_excel()
    excel_map = {**turbo_map, **tonghop_map}  # tonghop overrides turbo
    print(f'  Turbo file: {len(turbo_map)} mappings')
    print(f'  TongHop file: {len(tonghop_map)} mappings')
    print(f'  Total unique: {len(excel_map)} mappings')

    # Find products with numeric HangMay
    numeric_hms = HangMay.objects.all()
    numeric_hms = [h for h in numeric_hms if is_numeric(h.ten)]

    numeric_ids = set(h.id for h in numeric_hms)
    products_to_fix = Product.objects.filter(hang_may_id__in=numeric_ids, is_active=True)
    print(f'\nProducts with numeric HangMay: {products_to_fix.count()}')

    results = {
        'fixed_from_excel': 0,
        'fixed_from_name': 0,
        'moved_to_chua_ro': 0,
        'errors': 0,
    }

    chua_ro = get_or_create_hangmay('CHƯA RÕ')

    for p in products_to_fix:
        old_hm = p.hang_may.ten
        new_brand = None
        source = ''

        # Strategy 1: Excel mapping by ma_vt
        new_brand = excel_map.get(p.ma_vt)
        if new_brand:
            source = 'excel'

        # Strategy 2: Extract from ma_vt (may contain brand name)
        if not new_brand:
            new_brand = extract_brand_from_name(p.ma_vt)
            if new_brand:
                source = 'ma_vt'

        # Strategy 3: Extract from ten_hang
        if not new_brand:
            new_brand = extract_brand_from_name(p.ten_hang)
            if new_brand:
                source = 'ten_hang'

        # Strategy 4: Extract from model_turbo
        if not new_brand:
            new_brand = extract_brand_from_name(p.model_turbo)
            if new_brand:
                source = 'model_turbo'

        if new_brand and new_brand != old_hm:
            if not dry_run:
                new_hm = get_or_create_hangmay(new_brand)
                p.hang_may = new_hm
                p.save(update_fields=['hang_may', 'updated_at'])
            if source == 'excel':
                results['fixed_from_excel'] += 1
            else:
                results['fixed_from_name'] += 1
            print(f'  [{p.ma_vt}] {old_hm} → {new_brand} ({source})  ({p.ten_hang[:60]})')
        elif not new_brand:
            # Move to CHƯA RÕ
            if not dry_run:
                p.hang_may = chua_ro
                p.save(update_fields=['hang_may', 'updated_at'])
            results['moved_to_chua_ro'] += 1
            print(f'  [{p.ma_vt}] {old_hm} → CHƯA RÕ  ({p.ten_hang[:60]})')

    # Summary
    sep = '=' * 60
    print(f'\n{sep}')
    print('KET QUA')
    print(f'  Fixed from Excel mapping: {results["fixed_from_excel"]}')
    print(f'  Fixed from product name: {results["fixed_from_name"]}')
    print(f'  Moved to CHUA RO: {results["moved_to_chua_ro"]}')
    if dry_run:
        print(f'  [DRY RUN] - Chua ghi vao DB')
    print(sep)

    # Show numeric HangMay that will become empty
    remaining = HangMay.objects.filter(
        id__in=numeric_ids
    ).annotate(
        pc=models.Count('products')
    ).filter(pc=0) if not dry_run else []

    if dry_run:
        print('\nSau khi fix, cac HangMay numeric se bi xoa (ko con SP):')
        for h in sorted(numeric_hms, key=lambda h: -h.products.count())[:10]:
            print(f'  {h.ten}: {h.products.count()} SP')

    return results


if __name__ == '__main__':
    from django.db import models as dj_models
    parser = argparse.ArgumentParser(description='Fix HangMay numeric -> real brand')
    parser.add_argument('--dry-run', action='store_true', help='Preview only')
    args = parser.parse_args()
    fix_hangmay(dry_run=args.dry_run)
