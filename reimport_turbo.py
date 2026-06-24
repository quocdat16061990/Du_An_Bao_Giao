"""
Re-import Turbo & Ruot products from BAO_GIA_TURBO Excel with CORRECT column mapping.
- Deletes ALL existing turbo/ruot/so_linh_kien_turbo products
- Re-imports from Excel with proper HangMay, ThuongHieu, prices, tech specs

Usage:
    py reimport_turbo.py --dry-run    # Preview
    py reimport_turbo.py              # Run for real
"""

import sys, os, re, argparse
from pathlib import Path
from decimal import Decimal, InvalidOperation

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'backend.settings')
env_file = Path(__file__).resolve().parent / 'backend' / '.env'
if env_file.exists():
    for line in env_file.read_text(encoding='utf-8').splitlines():
        line = line.strip()
        if line and not line.startswith('#') and '=' in line:
            key, _, val = line.partition('=')
            os.environ.setdefault(key.strip(), val.strip())

import django; django.setup()
from products.models import Product, HangMay, HangSx, ThuongHieu, Category
from django.utils.text import slugify
import openpyxl


DOCS_DIR = Path(__file__).resolve().parent / 'docs'
EXCEL_FILE = DOCS_DIR / 'BAO_GIA_TURBO_v13_GIA_CO_ANH.xlsx'

# Column mapping for BAO_GIA_TURBO (1-indexed)
COL = {
    'hang_may': 1,      # A
    'ma_vt': 2,         # B
    'hang_sx': 3,       # C
    'model_turbo': 4,   # D
    'ma_dong_co': 5,    # E
    'oem_part_no': 6,   # F
    'dac_diem': 7,      # G
    'ung_dung': 8,      # H
    'ghi_chu': 9,       # I
    'thuong_hieu': 10,  # J
    'gia_ban': 11,      # K
    'gia_uu_dai': 12,   # L
    'gia_vip': 13,      # M
    'gia_dl_10': 14,    # N
    'cg_duoi': 15,      # O
    'cg_dinh': 16,      # P
    'cg_so': 17,         # Q
    'cl_duoi': 18,      # R
    'cl_dinh': 19,      # S
    'cl_so': 20,        # T
}

SHEET_LOAI_MAP = {
    '🚗 BÁO GIÁ TURBO ': 'turbo',
    'RUỘT TURBO': 'ruot',
    '🔩 SÒ & LINH KIỆN': 'so_linh_kien_turbo',
}


def parse_price(val) -> Decimal | None:
    """Parse price from Excel cell."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return Decimal(str(int(val)))
    text = str(val).strip()
    if not text or text.lower() in ('none', '—', '-', ''):
        return None
    # Remove VND symbol, commas, dots
    clean = re.sub(r'[₫đVND\s,]', '', text)
    # Handle formats like "5.500.000" or "5,500,000"
    if '.' in clean and ',' in clean:
        clean = clean.replace('.', '').replace(',', '.')
    elif '.' in clean:
        parts = clean.split('.')
        if len(parts) > 1 and len(parts[-1]) == 3:
            clean = clean.replace('.', '')
    elif ',' in clean:
        parts = clean.split(',')
        if len(parts) > 1 and len(parts[-1]) == 3:
            clean = clean.replace(',', '')
        else:
            clean = clean.replace(',', '.')
    try:
        return Decimal(clean)
    except InvalidOperation:
        return None


def parse_float(val) -> Decimal | None:
    """Parse float value from Excel."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return Decimal(str(val))
    try:
        return Decimal(str(val).strip())
    except InvalidOperation:
        return None


def get_or_create(model_class, ten: str, **extra):
    """Get or create a model instance by 'ten' field."""
    obj = model_class.objects.filter(ten__iexact=ten).first()
    if obj:
        return obj
    slug = slugify(ten)
    return model_class.objects.create(ten=ten, slug=slug, **extra)


def clean_thuong_hieu(raw: str) -> str | None:
    """Clean Thương Hiệu value from Excel."""
    r = raw.strip()
    if not r or r in ('None', '—', '-'):
        return None
    # Known real brands
    KNOWN = ['JRONE', 'TBS', 'VIDARIR', 'FIRE', 'EE', 'MX', 'GARRETT', 'SL',
             'ISUZU', 'MOBIS', 'SL UK', 'CHÍNH', 'DN-1197']
    r_upper = r.upper()
    for k in KNOWN:
        if k in r_upper:
            return k
    # If it's a part number (contains digits), skip
    if re.search(r'\d', r):
        return None
    return r[:100]


def reimport_turbo(dry_run=True):
    """Main re-import function."""
    if not EXCEL_FILE.exists():
        print(f'[ERROR] Excel file not found: {EXCEL_FILE}')
        return

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    # Count existing
    existing = Product.objects.filter(
        loai__in=['turbo', 'ruot', 'so_linh_kien_turbo']
    ).count()
    print(f'Existing turbo/ruot/slk products to DELETE: {existing}')

    if not dry_run:
        # Delete existing
        deleted, _ = Product.objects.filter(
            loai__in=['turbo', 'ruot', 'so_linh_kien_turbo']
        ).delete()
        print(f'Deleted: {deleted}')

    # Get or create categories
    turbo_cat = get_or_create(Category, ten='Turbo', mo_ta='Bộ turbo tăng áp đầy đủ')
    ruot_cat = get_or_create(Category, ten='Ruột Turbo', mo_ta='Ruột/Core turbo tăng áp')
    slk_cat = get_or_create(Category, ten='Sò & Linh Kiện Turbo', mo_ta='Sò lửa, sò gió & linh kiện turbo')

    loai_cat_map = {
        'turbo': turbo_cat,
        'ruot': ruot_cat,
        'so_linh_kien_turbo': slk_cat,
    }

    stats = {'created': 0, 'skipped': 0, 'errors': 0}

    for sname, loai in SHEET_LOAI_MAP.items():
        if sname not in wb.sheetnames:
            print(f'[WARN] Sheet not found: {sname}')
            continue

        ws = wb[sname]
        cat = loai_cat_map[loai]

        print(f'\n--- {sname} -> loai={loai} ---')

        for row in range(5, ws.max_row + 1):
            ma_vt = str(ws.cell(row, COL['ma_vt']).value or '').strip()

            # Skip non-HH rows (category separators, empty)
            if not ma_vt.startswith('HH'):
                continue

            hang_may_raw = str(ws.cell(row, COL['hang_may']).value or '').strip()
            # Clean └─ prefix
            hang_may_raw = hang_may_raw.replace('└─', '').strip()
            if not hang_may_raw or hang_may_raw in ('None', '—', ''):
                stats['skipped'] += 1
                continue

            hang_sx_raw = str(ws.cell(row, COL['hang_sx']).value or '').strip()
            thuong_hieu_raw = str(ws.cell(row, COL['thuong_hieu']).value or '').strip()
            model_turbo = str(ws.cell(row, COL['model_turbo']).value or '').strip()
            ma_dong_co = str(ws.cell(row, COL['ma_dong_co']).value or '').strip()
            oem_part_no = str(ws.cell(row, COL['oem_part_no']).value or '').strip()
            dac_diem = str(ws.cell(row, COL['dac_diem']).value or '').strip()
            ung_dung = str(ws.cell(row, COL['ung_dung']).value or '').strip()
            ghi_chu = str(ws.cell(row, COL['ghi_chu']).value or '').strip()

            # Prices
            gia_vip = parse_price(ws.cell(row, COL['gia_vip']).value)
            gia_uu_dai = parse_price(ws.cell(row, COL['gia_uu_dai']).value)
            gia_ban = parse_price(ws.cell(row, COL['gia_ban']).value)
            gia_dl_10 = parse_price(ws.cell(row, COL['gia_dl_10']).value)

            # Tech specs
            cg_duoi = parse_float(ws.cell(row, COL['cg_duoi']).value)
            cg_dinh = parse_float(ws.cell(row, COL['cg_dinh']).value)
            cg_so = str(ws.cell(row, COL['cg_so']).value or '').strip()
            cl_duoi = parse_float(ws.cell(row, COL['cl_duoi']).value)
            cl_dinh = parse_float(ws.cell(row, COL['cl_dinh']).value)
            cl_so = str(ws.cell(row, COL['cl_so']).value or '').strip()

            if dry_run:
                stats['created'] += 1
                if stats['created'] <= 3:
                    print(f'  [{ma_vt}] hm={hang_may_raw} sx={hang_sx_raw} th={thuong_hieu_raw}')
                    print(f'           model={model_turbo[:40]} dongco={ma_dong_co[:30]}')
                    print(f'           VIP={gia_vip} UU_DAI={gia_uu_dai} BAN={gia_ban} DL10={gia_dl_10}')
                continue

            # Get or create FK objects
            hm = get_or_create(HangMay, ten=hang_may_raw)
            hs = get_or_create(HangSx, ten=hang_sx_raw) if hang_sx_raw and hang_sx_raw not in ('None', '—', '') else None

            th = None
            cleaned_th = clean_thuong_hieu(thuong_hieu_raw)
            if cleaned_th:
                th = get_or_create(ThuongHieu, ten=cleaned_th)

            # Create product
            try:
                Product.objects.create(
                    ma_vt=ma_vt,
                    ten_hang=model_turbo,  # Model Turbo as primary name
                    model_turbo=model_turbo,
                    ma_dong_co=ma_dong_co or '',
                    oem_part_no=oem_part_no or '',
                    dac_diem=dac_diem or '',
                    ung_dung=ung_dung or '',
                    ghi_chu=ghi_chu or '',
                    loai=loai,
                    hang_may=hm,
                    hang_sx=hs,
                    thuong_hieu=th,
                    category=cat,
                    gia_vip=gia_vip,
                    gia_uu_dai=gia_uu_dai,
                    gia_dai_ly=gia_ban,  # GIÁ BÁN -> gia_dai_ly
                    gia_dl_10=gia_dl_10,
                    cg_duoi=cg_duoi,
                    cg_dinh=cg_dinh,
                    cg_so=cg_so or '',
                    cl_duoi=cl_duoi,
                    cl_dinh=cl_dinh,
                    cl_so=cl_so or '',
                    sheet_name=sname.strip(),
                    is_active=True,
                )
                stats['created'] += 1
            except Exception as e:
                stats['errors'] += 1
                if stats['errors'] <= 5:
                    print(f'  ERROR [{ma_vt}]: {e}')

    wb.close()

    # Summary
    sep = '=' * 60
    print(f'\n{sep}')
    print('KET QUA')
    print(f'  Created: {stats["created"]}')
    print(f'  Skipped: {stats["skipped"]}')
    print(f'  Errors: {stats["errors"]}')
    if dry_run:
        print(f'  [DRY RUN] - Chua ghi vao DB')
    else:
        total = Product.objects.filter(loai__in=['turbo', 'ruot', 'so_linh_kien_turbo']).count()
        print(f'  Total turbo/ruot/slk in DB: {total}')
    print(sep)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Re-import Turbo products from Excel')
    parser.add_argument('--dry-run', action='store_true', help='Preview only')
    args = parser.parse_args()
    reimport_turbo(dry_run=args.dry_run)
