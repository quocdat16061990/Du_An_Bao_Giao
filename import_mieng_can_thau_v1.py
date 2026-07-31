"""
Import miểng, căn dọc, bạc thau từ TRA CỨU MIỂNG _CĂN _THAU.xlsx
- Đưa vào các danh mục tương ứng
- Thuộc tính kỹ thuật lưu vào JSON attributes
- Backs up existing images by ma_vt
- Deletes old data, imports new data
"""

import sys, os, re, argparse, math
from pathlib import Path

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'backend.settings')
env_file = Path(__file__).resolve().parent / 'backend' / '.env'
if env_file.exists():
    for line in env_file.read_text(encoding='utf-8').splitlines():
        line = line.strip()
        if line and not line.startswith('#') and '=' in line:
            key, _, val = line.partition('=')
            os.environ.setdefault(key.strip(), val.strip())

import django; django.setup()
from products.models import Product, HangMay, Category
from django.utils.text import slugify
from django.db import transaction
import openpyxl

DOCS_DIR = Path(__file__).resolve().parent / 'docs' / 'update-docs-v1'
EXCEL_FILE = DOCS_DIR / 'TRA CỨU MIỂNG _CĂN _THAU.xlsx'

def get_or_create(model_class, ten: str):
    if not ten or ten in ('None', '—', '', ' '): return None
    obj = model_class.objects.filter(ten__iexact=ten).first()
    if obj: return obj
    base_slug = slugify(ten)
    if not base_slug: base_slug = 'khac'

    import time
    try:
        obj, _ = model_class.objects.get_or_create(ten=ten, defaults={'slug': f'{base_slug}-{int(time.time()*1000)}'})
        return obj
    except Exception:
        return None

def is_nan(val):
    if val is None: return True
    if isinstance(val, float) and math.isnan(val): return True
    return False

def clean_val(val):
    if is_nan(val): return ''
    s = str(val).strip()
    if s in ('None', 'NaN', 'nan', '—', '-'): return ''
    return s

def map_loai(loai_bac_str):
    s = clean_val(loai_bac_str).lower()
    if 'thrust' in s or 'căn dọc' in s: return 'can_thau'
    return 'mieng_bac'

def process_general_sheet(ws, start_row, parse_func, stats, image_map):
    current_hang = None
    for row in range(start_row, ws.max_row + 1):
        data = parse_func(ws, row)
        if not data: continue

        # Inherit hãng máy from previous row if current is empty
        if data.get('hang_may_raw'):
            current_hang = data['hang_may_raw']
        elif current_hang:
            data['hang_may_raw'] = current_hang

        ma_vt = data.get('ma_vt')
        if not ma_vt: continue

        ma_vt = str(ma_vt)[:100]
        hm_name = str(data.get('hang_may_raw') or 'Khác')[:100]

        if data.get('dry_run'):
            stats['created'] += 1
            continue

        hm = get_or_create(HangMay, ten=hm_name)
        if not hm: hm = get_or_create(HangMay, ten='Khác')

        try:
            obj = Product.objects.create(
                ma_vt=ma_vt,
                loai=data.get('loai', 'mieng_bac'),
                ten_hang=data.get('ten_hang', '')[:500],
                hang_may=hm,
                parno=data.get('parno', '')[:300],
                is_active=True,
                attributes=data.get('attributes', {})
            )
            stats['created'] += 1
            if ma_vt in image_map:
                obj.hinh_anh = image_map[ma_vt]['hinh_anh']
                obj.danh_sach_hinh_anh = image_map[ma_vt]['danh_sach_hinh_anh']
                obj.save(update_fields=['hinh_anh', 'danh_sach_hinh_anh'])
                stats['restored_images'] += 1
        except Exception as e:
            stats['errors'] += 1
            if stats['errors'] <= 3:
                print(f"  ERROR [{ma_vt}]: {e}")


@transaction.atomic
def import_mieng(dry_run=True):
    if not EXCEL_FILE.exists():
        print(f'[ERROR] File not found: {EXCEL_FILE}')
        return

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    types_to_process = ['mieng_bac', 'can_thau']

    # Phase 1: Backup images
    print("--- Backing up images ---")
    image_map = {}
    products_to_delete = Product.objects.filter(loai__in=types_to_process)
    for p in products_to_delete:
        if p.hinh_anh or p.danh_sach_hinh_anh:
            image_map[p.ma_vt] = {
                'hinh_anh': p.hinh_anh,
                'danh_sach_hinh_anh': p.danh_sach_hinh_anh
            }
    print(f"Backed up images for {len(image_map)} products.")

    # Phase 2: Delete existing
    print(f'Products to DELETE ({len(types_to_process)} types): {products_to_delete.count()}')
    if not dry_run:
        deleted, _ = products_to_delete.delete()
        print(f'Deleted: {deleted}')

    stats = {'created': 0, 'restored_images': 0, 'skipped': 0, 'errors': 0}

    # 1. Sheet MIỂNG
    if 'MIỂNG' in wb.sheetnames:
        print("\n--- Sheet: MIỂNG ---")
        ws = wb['MIỂNG']

        def parse_mieng(ws, row):
            hm = clean_val(ws.cell(row, 1).value)
            if hm.startswith('◆'): return None
            parno = clean_val(ws.cell(row, 5).value)
            if not parno: return None
            return {
                'dry_run': dry_run,
                'hang_may_raw': hm,
                'ma_vt': parno,
                'parno': parno,
                'ten_hang': clean_val(ws.cell(row, 4).value) + ' ' + clean_val(ws.cell(row, 2).value),
                'loai': 'mieng_bac',
                'attributes': {
                    'ma_dong_co': clean_val(ws.cell(row, 2).value),
                    'loai_chi_tiet': clean_val(ws.cell(row, 3).value),
                    'cung_trong': clean_val(ws.cell(row, 6).value),
                    'cung_ngoai': clean_val(ws.cell(row, 7).value),
                    'chieu_cao': clean_val(ws.cell(row, 8).value),
                    'chieu_day': clean_val(ws.cell(row, 9).value),
                    'hang_sx': clean_val(ws.cell(row, 11).value)
                }
            }
        process_general_sheet(ws, 4, parse_mieng, stats, image_map)

    # 2. Sheet CĂN DỌC
    if 'CĂN DỌC' in wb.sheetnames:
        print("\n--- Sheet: CĂN DỌC ---")
        ws = wb['CĂN DỌC']

        def parse_can(ws, row):
            hm = clean_val(ws.cell(row, 1).value)
            if hm.startswith('◆'): return None
            parno = clean_val(ws.cell(row, 4).value)
            if not parno: return None
            return {
                'dry_run': dry_run,
                'hang_may_raw': hm,
                'ma_vt': parno,
                'parno': parno,
                'ten_hang': clean_val(ws.cell(row, 3).value) + ' ' + clean_val(ws.cell(row, 2).value),
                'loai': 'can_thau',
                'attributes': {
                    'ma_dong_co': clean_val(ws.cell(row, 2).value),
                    'od': clean_val(ws.cell(row, 5).value),
                    'id': clean_val(ws.cell(row, 6).value),
                    'day': clean_val(ws.cell(row, 7).value),
                    'so_mieng': clean_val(ws.cell(row, 8).value)
                }
            }
        process_general_sheet(ws, 5, parse_can, stats, image_map)

    # 3. Sheets Brand (TAIHO, NDC, DAIDO)
    for sname in ['TAIHO', 'NDC', 'DAIDO']:
        if sname in wb.sheetnames:
            print(f"\n--- Sheet: {sname} ---")
            ws = wb[sname]

            def parse_brand(ws, row):
                hm = clean_val(ws.cell(row, 1).value)
                loai_bac_str = clean_val(ws.cell(row, 3).value)
                set_no = clean_val(ws.cell(row, 4).value)
                if not set_no: return None
                return {
                    'dry_run': dry_run,
                    'hang_may_raw': hm,
                    'ma_vt': set_no,
                    'parno': clean_val(ws.cell(row, 5).value),  # Part no
                    'ten_hang': f"[{sname}] {loai_bac_str} {clean_val(ws.cell(row, 2).value)}",
                    'loai': map_loai(loai_bac_str),
                    'attributes': {
                        'brand': sname,
                        'model': clean_val(ws.cell(row, 2).value),
                        'shaft_dia': clean_val(ws.cell(row, 6).value),
                        'housing_dia': clean_val(ws.cell(row, 7).value),
                        'day': clean_val(ws.cell(row, 8).value),
                        'dai': clean_val(ws.cell(row, 9).value),
                        'ref_no': clean_val(ws.cell(row, 10).value)
                    }
                }
            process_general_sheet(ws, 2, parse_brand, stats, image_map)

    wb.close()
    print(f"\nImport finished. Stats: {stats}")

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--dry-run', action='store_true', help='Run without modifying DB')
    args = parser.parse_args()

    import_mieng(dry_run=args.dry_run)
