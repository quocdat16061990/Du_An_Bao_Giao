from django.test import TestCase
from django.urls import reverse
from rest_framework import status
from rest_framework.test import APIClient
from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from decimal import Decimal
from .models import Product, Category, HangMay, HangSx, ThuongHieu, Customer


class SmartSearchTestCase(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.hang_may, _ = HangMay.objects.get_or_create(ten="TOYOTA", slug="toyota")
        self.cat_xy_lanh, _ = Category.objects.get_or_create(ten="Xy lanh", slug="xy-lanh")
        self.cat_piston, _ = Category.objects.get_or_create(ten="Piston", slug="piston")

        # Tạo sản phẩm 1: Động cơ 6D125, Piston
        self.p1 = Product.objects.create(
            ma_vt="HH01001",
            ten_hang="Piston 6D125",
            model_turbo="CT16",
            ma_dong_co="6D125",
            hang_may=self.hang_may,
            category=self.cat_piston,
            loai="piston",
            gia_von=Decimal("5000000"),
            gia_vip=Decimal("6000000"),
            gia_uu_dai=Decimal("6200000"),
            gia_dai_ly=Decimal("6500000"),
            gia_gara=Decimal("6800000"),
            gia_dl_10=Decimal("7000000"),
        )

        # Tạo sản phẩm 2: Động cơ 6D125, Xy lanh
        self.p2 = Product.objects.create(
            ma_vt="HH01002",
            ten_hang="Xy lanh 6D125",
            model_turbo="CT16V",
            ma_dong_co="6D125",
            hang_may=self.hang_may,
            category=self.cat_xy_lanh,
            loai="xy_lanh",
            gia_von=Decimal("3000000"),
            gia_vip=Decimal("3500000"),
            gia_uu_dai=Decimal("3800000"),
            gia_dai_ly=Decimal("4000000"),
            gia_gara=Decimal("4200000"),
            gia_dl_10=Decimal("4500000"),
        )

        # Tạo sản phẩm 3: Động cơ 4D30, Piston
        self.p3 = Product.objects.create(
            ma_vt="HH01003",
            ten_hang="Piston 4D30",
            model_turbo="",
            ma_dong_co="4D30",
            hang_may=self.hang_may,
            category=self.cat_piston,
            loai="piston",
            gia_von=Decimal("2000000"),
            gia_vip=Decimal("2200000"),
            gia_uu_dai=Decimal("2400000"),
            gia_dai_ly=Decimal("2600000"),
            gia_gara=Decimal("2800000"),
            gia_dl_10=Decimal("3000000"),
        )

    def test_search_single_keyword(self):
        # Tìm kiếm 1 từ: 6D125 -> Phải ra p1 và p2
        response = self.client.get(reverse('product-list'), {'q': '6D125'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        results = response.data['results']
        self.assertEqual(len(results), 2)
        ids = [r['id'] for r in results]
        self.assertIn(self.p1.id, ids)
        self.assertIn(self.p2.id, ids)

    def test_search_multi_keyword_and_logic(self):
        # Tìm kiếm đa từ: '6d125 xy' -> Chỉ ra p2 (Xy lanh 6D125)
        response = self.client.get(reverse('product-list'), {'q': '6d125 xy'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        results = response.data['results']
        self.assertEqual(len(results), 1)
        self.assertEqual(results[0]['id'], self.p2.id)

    def test_search_multi_keyword_comma_separated(self):
        # Tìm kiếm đa từ phân tách bằng dấu phẩy: '6d125, pis' -> Chỉ ra p1
        response = self.client.get(reverse('product-list'), {'q': '6d125, pis'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        results = response.data['results']
        self.assertEqual(len(results), 1)
        self.assertEqual(results[0]['id'], self.p1.id)

    def test_filter_price_type_gara_and_von(self):
        # Lọc các sản phẩm có gia_gara không null
        response = self.client.get(reverse('product-list'), {'phan_loai_gia': 'gara'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 3)

        # Lọc các sản phẩm có gia_von không null
        response = self.client.get(reverse('product-list'), {'phan_loai_gia': 'von'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 3)


class CustomerPriceTestCase(TestCase):
    def setUp(self):
        self.client = APIClient()
        # Tạo test user và force authenticate vì POST API yêu cầu đăng nhập
        self.user = User.objects.create_user(username='testadmin', password='testpassword')
        self.client.force_authenticate(user=self.user)

        self.hang_may, _ = HangMay.objects.get_or_create(ten="TOYOTA", slug="toyota")
        self.p1 = Product.objects.create(
            ma_vt="HH02001",
            ten_hang="Turbo TD04",
            model_turbo="TD04",
            ma_dong_co="4D56",
            hang_may=self.hang_may,
            loai="turbo",
            gia_von=Decimal("4000000"),
            gia_vip=Decimal("4500000"),
            gia_uu_dai=Decimal("4800000"),
            gia_dai_ly=Decimal("5000000"),
            gia_gara=Decimal("5200000"),
            gia_dl_10=Decimal("5500000"),
        )

    def test_create_customer_with_new_phan_loai(self):
        # Tạo khách hàng Gara
        response = self.client.post(reverse('customer-list'), {
            "ten_kh": "Gara Autotech",
            "dien_thoai": "0987654321",
            "phan_loai": "GARA",
            "dia_chi": "Hà Nội"
        })
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data['phan_loai'], "GARA")

        # Tạo khách hàng Đại lý
        response = self.client.post(reverse('customer-list'), {
            "ten_kh": "Đại lý Toàn Cầu",
            "dien_thoai": "0123456789",
            "phan_loai": "ĐẠI_LÝ",
            "dia_chi": "Hồ Chí Minh"
        })
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data['phan_loai'], "ĐẠI_LÝ")

    def test_product_price_mapping_for_gara_and_dai_ly(self):
        # Kiểm tra ánh xạ trực tiếp trong Product model
        self.assertEqual(self.p1.get_price_for_type("GARA"), Decimal("5200000"))
        self.assertEqual(self.p1.get_price_for_type("ĐẠI_LÝ"), Decimal("5000000"))

        # Gửi thử tạo báo giá nháp cho khách Gara
        customer = Customer.objects.create(
            ma_kh="KH001",
            ten_kh="Gara Autotech",
            phan_loai="GARA"
        )
        response = self.client.post(reverse('quotation-preview'), {
            "customer_id": customer.id,
            "product_ids": [self.p1.id]
        })
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        # Đơn giá phải là 5,200,000 đ
        product_quote = response.data['products'][0]
        self.assertEqual(int(float(product_quote['don_gia'])), 5200000)


class QuotationExportExcelTestCase(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(username='quote_admin', password='password123')
        self.client.force_authenticate(user=self.user)

        self.hang_may, _ = HangMay.objects.get_or_create(ten="ISUZU", slug="isuzu")

        self.customer = Customer.objects.create(
            ma_kh="KH_QT_01",
            ten_kh="Công Ty Anh Nguyên",
            dien_thoai="0961046767",
            dia_chi="Sài Gòn",
            phan_loai="ĐẠI_LÝ",
        )

        self.p1 = Product.objects.create(
            ma_vt="HH070412",
            ten_hang="Piston 6BD1 nổ vuông dài 104",
            dvt="BỘ 6",
            hang_may=self.hang_may,
            gia_dai_ly=Decimal("5000000"),
        )
        self.p2 = Product.objects.create(
            ma_vt="HH083107",
            ten_hang="Séc măng 6BD1 3-2.5-5",
            dvt="BỘ 6",
            hang_may=self.hang_may,
            gia_dai_ly=Decimal("700000"),
        )

    def test_build_quotation_excel_structure(self):
        from io import BytesIO
        from openpyxl import load_workbook
        from .quotation_excel import build_quotation_excel

        excel_bytes = build_quotation_excel(self.customer, [self.p1, self.p2], "BG20260731-01")
        self.assertTrue(len(excel_bytes) > 0)

        wb = load_workbook(BytesIO(excel_bytes), data_only=False)
        ws = wb.active

        # Check headers
        self.assertIn("CÔNG TY TNHH MÁY CÔNG TRÌNH MIỀN NAM", str(ws['A1'].value))
        self.assertIn("Anh Nguyên", str(ws['A9'].value))

        # Check column headers (Row 14)
        self.assertEqual(ws['A14'].value, "STT")
        self.assertEqual(ws['B14'].value, "TÊN HÀNG HÓA")
        self.assertEqual(ws['C14'].value, "MÃ HH")
        self.assertEqual(ws['D14'].value, "ĐVT")
        self.assertEqual(ws['E14'].value, "SL")
        self.assertEqual(ws['F14'].value, "ĐƠN GIÁ")
        self.assertEqual(ws['G14'].value, "THÀNH TIỀN")

        # Check product rows
        self.assertEqual(ws['A15'].value, 1)
        self.assertEqual(ws['B15'].value, "Piston 6BD1 nổ vuông dài 104")
        self.assertEqual(ws['C15'].value, "HH070412")
        self.assertEqual(ws['F15'].value, 5000000)
        self.assertEqual(ws['G15'].value, "=F15*E15")

        self.assertEqual(ws['A16'].value, 2)
        self.assertEqual(ws['B16'].value, "Séc măng 6BD1 3-2.5-5")
        self.assertEqual(ws['C16'].value, "HH083107")

        # Check summary rows
        self.assertEqual(ws['A20'].value, "Cộng tiền hàng (chưa VAT)")
        self.assertEqual(ws['G20'].value, "=ROUND(G22/1.08,0)")

        self.assertEqual(ws['A21'].value, "Thuế GTGT 8%")
        self.assertEqual(ws['G21'].value, "=G22-G20")

        self.assertEqual(ws['A22'].value, "TỔNG CỘNG (đã có VAT)")
        self.assertEqual(ws['G22'].value, "=SUM(G15:G19)")

        self.assertIn("Bằng chữ:", str(ws['A23'].value))

    def test_quotation_export_excel_api(self):
        response = self.client.post(reverse('quotation-export-excel'), {
            "customer_id": self.customer.id,
            "product_ids": [self.p1.id, self.p2.id],
        })
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertTrue(len(response.content) > 0)


class OrderExportExcelTestCase(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(username='order_admin', password='password123')
        self.client.force_authenticate(user=self.user)

        self.hang_may, _ = HangMay.objects.get_or_create(ten="ISUZU", slug="isuzu")

        self.customer = Customer.objects.create(
            ma_kh="KH_OD_01",
            ten_kh="Gara Hoàng Long",
            dien_thoai="0912345678",
            dia_chi="Hà Nội",
            phan_loai="GARA",
        )

        self.p1 = Product.objects.create(
            ma_vt="HH001",
            ten_hang="Bộ Ron Isuzu 6BD1",
            dvt="Bộ",
            hang_may=self.hang_may,
            gia_gara=Decimal("1030000"),
        )
        self.p2 = Product.objects.create(
            ma_vt="HH002",
            ten_hang="Xy lanh 6BD1 kiếng",
            dvt="BỘ 6",
            hang_may=self.hang_may,
            gia_gara=Decimal("2800000"),
        )

    def test_build_order_excel_structure(self):
        from io import BytesIO
        from openpyxl import load_workbook
        from .order_excel import build_order_excel

        excel_bytes = build_order_excel(self.customer, [self.p1, self.p2], "DH20260731-01")
        self.assertTrue(len(excel_bytes) > 0)

        wb = load_workbook(BytesIO(excel_bytes), data_only=False)
        ws = wb.active

        # Check customer info
        self.assertEqual(ws['C3'].value, "Gara Hoàng Long")
        self.assertEqual(ws['C4'].value, "0912345678")

        # Check table headers (Row 8)
        self.assertEqual(ws['A8'].value, "STT")
        self.assertEqual(ws['B8'].value, "NGÀY ĐẶT")
        self.assertEqual(ws['C8'].value, "MÃ HH")
        self.assertEqual(ws['D8'].value, "TÊN SẢN PHẨM")
        self.assertEqual(ws['E8'].value, "ĐVT")
        self.assertEqual(ws['F8'].value, "SỐ LƯỢNG")
        self.assertEqual(ws['G8'].value, "ĐƠN GIÁ")
        self.assertEqual(ws['H8'].value, "THÀNH TIỀN")

        # Check line items
        self.assertEqual(ws['A9'].value, 1)
        self.assertEqual(ws['C9'].value, "HH001")
        self.assertEqual(ws['D9'].value, "Bộ Ron Isuzu 6BD1")
        self.assertEqual(ws['G9'].value, 1030000)

        # Check summary totals
        self.assertEqual(ws['A24'].value, "TỔNG TIỀN HÀNG")
        self.assertEqual(ws['H24'].value, "=SUM(H9:H23)")

        self.assertEqual(ws['A25'].value, "VAT 8%")
        self.assertEqual(ws['H25'].value, "=H24*8%")

        self.assertEqual(ws['A26'].value, "TỔNG THANH TOÁN")
        self.assertEqual(ws['H26'].value, "=H24+H25")

    def test_order_export_excel_api(self):
        response = self.client.post(reverse('order-export-excel'), {
            "customer_id": self.customer.id,
            "product_ids": [self.p1.id, self.p2.id],
        })
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertTrue(len(response.content) > 0)


class ProductUpdateSerializerTestCase(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(username='admin', password='password123')
        self.client.force_authenticate(user=self.user)

        self.hang_may1, _ = HangMay.objects.get_or_create(ten="MITSUBISHI", slug="mitsubishi")
        self.hang_may2, _ = HangMay.objects.get_or_create(ten="KOMATSU", slug="komatsu")
        self.hang_sx, _ = HangSx.objects.get_or_create(ten="Mitsubishi", slug="mitsubishi-sx")
        self.thuong_hieu, _ = ThuongHieu.objects.get_or_create(ten="SL", slug="sl")
        self.category, _ = Category.objects.get_or_create(ten="Ruột Turbo", slug="ruot-turbo")

        self.product = Product.objects.create(
            ma_vt="HH080019",
            ten_hang="Ruột Turbo S6R",
            loai="ruot",
            hang_may=self.hang_may1,
            hang_sx=self.hang_sx,
            thuong_hieu=self.thuong_hieu,
            category=self.category,
            gia_vip=Decimal("7800000"),
            gia_uu_dai=Decimal("7500000"),
            gia_dl_10=Decimal("8580000"),
        )

    def test_patch_product_with_nested_dicts(self):
        url = reverse('product-detail', kwargs={'pk': self.product.id})
        payload = {
            "loai": "ruot",
            "ma_vt": "HH080019",
            "hang_may": {"id": self.hang_may2.id, "ten": "KOMATSU", "slug": "komatsu"},
            "hang_sx": {"id": self.hang_sx.id, "ten": "Mitsubishi"},
            "thuong_hieu": {"id": self.thuong_hieu.id, "ten": "SL"},
            "category": {"id": self.category.id, "ten": "Ruột Turbo"},
            "gia_vip": 8000000,
            "ghi_chu": "Cập nhật với nested dict"
        }
        response = self.client.patch(url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        self.product.refresh_from_db()
        self.assertEqual(self.product.hang_may, self.hang_may2)
        self.assertEqual(int(self.product.gia_vip), 8000000)
        self.assertEqual(self.product.ghi_chu, "Cập nhật với nested dict")

    def test_patch_product_with_integer_pks(self):
        url = reverse('product-detail', kwargs={'pk': self.product.id})
        payload = {
            "hang_may": self.hang_may2.id,
            "gia_vip": 8200000
        }
        response = self.client.patch(url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        self.product.refresh_from_db()
        self.assertEqual(self.product.hang_may, self.hang_may2)
        self.assertEqual(int(self.product.gia_vip), 8200000)

    def test_patch_product_set_null_optional_fields(self):
        url = reverse('product-detail', kwargs={'pk': self.product.id})
        payload = {
            "hang_sx": None,
            "thuong_hieu": None,
            "category": None
        }
        response = self.client.patch(url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        self.product.refresh_from_db()
        self.assertIsNone(self.product.hang_sx)
        self.assertIsNone(self.product.thuong_hieu)
        self.assertIsNone(self.product.category)

    def test_patch_product_invalid_dict_without_id(self):
        url = reverse('product-detail', kwargs={'pk': self.product.id})
        payload = {
            "hang_may": {"ten": "Invalid Dict"}
        }
        response = self.client.patch(url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("hang_may", response.data)

    def test_patch_product_non_existent_pk(self):
        url = reverse('product-detail', kwargs={'pk': self.product.id})
        payload = {
            "hang_may": 999999
        }
        response = self.client.patch(url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("hang_may", response.data)

    def test_patch_single_field_leaves_others_intact(self):
        url = reverse('product-detail', kwargs={'pk': self.product.id})
        payload = {
            "ghi_chu": "Chỉ cập nhật ghi chú mới"
        }
        response = self.client.patch(url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        self.product.refresh_from_db()
        self.assertEqual(self.product.ghi_chu, "Chỉ cập nhật ghi chú mới")
        # Ensure other fields remain completely intact
        self.assertEqual(self.product.ma_vt, "HH080019")
        self.assertEqual(self.product.hang_may, self.hang_may1)
        self.assertEqual(self.product.hang_sx, self.hang_sx)
        self.assertEqual(self.product.thuong_hieu, self.thuong_hieu)
        self.assertEqual(self.product.category, self.category)
        self.assertEqual(int(self.product.gia_vip), 7800000)


class ProductImageUploadTestCase(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(username='upload_admin', password='password123')
        self.client.force_authenticate(user=self.user)
        self.url = reverse('product-upload-image')

    def test_upload_jpg_image_success(self):
        file_obj = SimpleUploadedFile("test_turbo.jpg", b"fake_jpg_content", content_type="image/jpeg")
        response = self.client.post(self.url, {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertIn('url', response.data)
        self.assertEqual(response.data['name'], "test_turbo.jpg")
        self.assertTrue(response.data['url'].endswith('.jpg'))

    def test_upload_uppercase_jpg_success(self):
        file_obj = SimpleUploadedFile("20260407153611.JPG", b"fake_jpg_content", content_type="image/jpeg")
        response = self.client.post(self.url, {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertIn('url', response.data)
        self.assertEqual(response.data['name'], "20260407153611.JPG")
        self.assertTrue(response.data['url'].endswith('.jpg'))

    def test_upload_png_image_success(self):
        file_obj = SimpleUploadedFile("sample.png", b"fake_png_content", content_type="image/png")
        response = self.client.post(self.url, {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertTrue(response.data['url'].endswith('.png'))

    def test_upload_invalid_extension_fails(self):
        file_obj = SimpleUploadedFile("document.pdf", b"fake_pdf_content", content_type="application/pdf")
        response = self.client.post(self.url, {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('error', response.data)

    def test_upload_no_file_fails(self):
        response = self.client.post(self.url, {}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('error', response.data)
