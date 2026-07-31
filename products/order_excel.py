from copy import copy
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
import re

from django.conf import settings
from openpyxl import load_workbook
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.utils import get_column_letter, range_boundaries

TEMPLATE_PATH = Path(settings.BASE_DIR) / 'templates' / 'PHIEU_DAT_HANG_KHÁCH HÀNG.xlsx'
PRODUCT_START_ROW = 9
PRODUCT_TEMPLATE_LAST_ROW = 23
TOTAL_ROW = 24  # TỔNG TIỀN HÀNG row index in template


def _shift_merged_ranges(ws, start_row: int, amount: int):
    shifted = []
    for merged_range in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        if min_row >= start_row:
            ws.unmerge_cells(str(merged_range))
            shifted.append((min_col, min_row + amount, max_col, max_row + amount))

    for min_col, min_row, max_col, max_row in shifted:
        ws.merge_cells(
            f'{get_column_letter(min_col)}{min_row}:'
            f'{get_column_letter(max_col)}{max_row}'
        )


def _copy_row_style(ws, source_row: int, target_row: int):
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.number_format = source.number_format
            target.protection = copy(source.protection)


def _configure_print_area(ws):
    ws.print_area = f'A1:I{ws.max_row}'
    ws.sheet_properties.pageSetUpPr = ws.sheet_properties.pageSetUpPr or PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4


def build_order_excel(customer, products_qs, quote_number: str, custom_prices_map=None, creator_name=None) -> bytes:
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # Set sheet name to quote_number or standard
    ws.title = quote_number[:30]

    products = list(products_qs)
    total_template_rows = PRODUCT_TEMPLATE_LAST_ROW - PRODUCT_START_ROW + 1  # 15 rows
    extra_rows = max(0, len(products) - total_template_rows)

    if extra_rows > 0:
        _shift_merged_ranges(ws, TOTAL_ROW, extra_rows)
        ws.insert_rows(TOTAL_ROW, extra_rows)
        for row in range(TOTAL_ROW, TOTAL_ROW + extra_rows):
            _copy_row_style(ws, PRODUCT_TEMPLATE_LAST_ROW, row)

    total_row = TOTAL_ROW + extra_rows

    # Populate customer details
    ws['C3'] = customer.ten_kh
    ws['C4'] = customer.dien_thoai or ''
    ws['C5'] = customer.dia_chi or customer.tinh_tp or ''

    today_str = datetime.now().strftime('%d/%m/%Y')

    for index, product in enumerate(products, start=1):
        row = PRODUCT_START_ROW + index - 1
        qty = getattr(product, 'quantity', 1)

        if custom_prices_map and product.id in custom_prices_map:
            unit_price = Decimal(str(custom_prices_map[product.id]['price']))
            qty = custom_prices_map[product.id].get('quantity', 1)
        else:
            unit_price = product.get_price_for_type(customer.phan_loai) or Decimal('0')

        # Col A: STT
        ws.cell(row, 1).value = index
        # Col B: Ngày đặt
        ws.cell(row, 2).value = today_str
        # Col C: Mã HH
        ws.cell(row, 3).value = product.ma_vt
        # Col D: Tên sản phẩm
        ws.cell(row, 4).value = product.ten_hang or product.model_turbo or ''
        # Col E: ĐVT
        ws.cell(row, 5).value = product.dvt or 'Cái'
        # Col F: Số lượng
        ws.cell(row, 6).value = qty
        # Col G: Đơn giá
        ws.cell(row, 7).value = int(unit_price)
        # Col H: Thành tiền (Formula)
        ws.cell(row, 8).value = f'=IF(AND(F{row}<>"",G{row}<>""),F{row}*G{row},"")'
        # Col I: Ghi chú
        ws.cell(row, 9).value = getattr(product, 'ghi_chu', '') or ''
        # Col K: Copy Zalo formula
        ws.cell(row, 11).value = (
            f'=IF(D{row}="","",A{row}&". "&D{row}&IF(E{row}=""," "," ("&E{row}&") ")&'
            f'"| SL "&IF(F{row}="","?",SUBSTITUTE(TEXT(F{row},"#,##0"),",","."))&'
            f'IF(G{row}=""," | (chưa có giá)"," | "&SUBSTITUTE(TEXT(G{row},"#,##0"),",",".")&"đ")&'
            f'IF(H{row}=""," "," = "&SUBSTITUTE(TEXT(H{row},"#,##0"),",",".")&"đ"))'
        )

    # For empty template rows, clear the default cells but keep Col H and K formulas
    for row in range(PRODUCT_START_ROW + len(products), total_row):
        ws.cell(row, 1).value = None
        for col in [2, 3, 4, 5, 6, 7, 9]:
            ws.cell(row, col).value = None

    # Update summary row formulas (Row 7)
    last_item_row = total_row - 1
    ws['A7'] = f'=COUNTIF(D9:D{last_item_row},"?*")'
    ws['C7'] = f'=SUM(F9:F{last_item_row})'
    ws['E7'] = f'=SUMPRODUCT((D9:D{last_item_row}<>"")*(G9:G{last_item_row}=""))'
    ws['G7'] = f'=H{total_row + 2}'  # Points to TỔNG THANH TOÁN

    # Update total rows (TỔNG TIỀN HÀNG, VAT 8%, TỔNG THANH TOÁN)
    ws.cell(total_row, 1).value = 'TỔNG TIỀN HÀNG'
    ws.cell(total_row, 8).value = f'=SUM(H9:H{last_item_row})'

    ws.cell(total_row + 1, 1).value = 'VAT 8%'
    ws.cell(total_row + 1, 8).value = f'=H{total_row}*8%'

    ws.cell(total_row + 2, 1).value = 'TỔNG THANH TOÁN'
    ws.cell(total_row + 2, 8).value = f'=H{total_row}+H{total_row + 1}'

    # Update creator name at G32 + extra_rows
    creator_row = 32 + extra_rows
    ws.cell(creator_row, 7).value = creator_name or 'Nguyễn Văn Luân'

    _configure_print_area(ws)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
