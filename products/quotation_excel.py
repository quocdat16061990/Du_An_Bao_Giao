from copy import copy
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
import re

from django.conf import settings
from openpyxl import load_workbook
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.utils import get_column_letter, range_boundaries


TEMPLATE_PATH = Path(settings.BASE_DIR) / 'templates' / 'bao_gia_template_clean.xlsx'
PRODUCT_START_ROW = 13
PRODUCT_TEMPLATE_LAST_ROW = 17
TOTAL_ROW = 18
VAT_RATE = Decimal('0.08')


PRICE_LABELS = {
    'VIP': 'Giá VIP',
    'UU_DAI': 'Giá ưu đãi',
    'ƯU_ĐÃI': 'Giá ưu đãi',
    'DAI_LY': 'Giá đại lý',
    'ĐẠI_LÝ': 'Giá đại lý',
    'GARA': 'Giá gara',
    'NGOAI_LE': 'Giá DL+10%',
    'NGOẠI_LỆ': 'Giá DL+10%',
}


def _shift_merged_ranges(ws, start_row: int, amount: int):
    shifted = []
    for merged_range in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        if min_row >= start_row:
            ws.unmerge_cells(str(merged_range))
            shifted.append((min_col, min_row + amount, max_col, max_row + amount))

    for min_col, min_row, max_col, max_row in shifted:
        ws.merge_cells(
            f'{get_column_letter(min_col)}{min_row}:'
            f'{get_column_letter(max_col)}{max_row}'
        )


def _copy_row_style(ws, source_row: int, target_row: int):
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.number_format = source.number_format
            target.protection = copy(source.protection)


def _configure_print_area(ws):
    ws.print_area = f'A1:G{ws.max_row}'
    ws.sheet_properties.pageSetUpPr = ws.sheet_properties.pageSetUpPr or PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4


def _digits_to_words(number: int) -> str:
    names = [
        'không', 'một', 'hai', 'ba', 'bốn',
        'năm', 'sáu', 'bảy', 'tám', 'chín',
    ]
    if number == 0:
        return 'không'

    def read_three(value: int, full: bool) -> str:
        hundred = value // 100
        ten = (value % 100) // 10
        unit = value % 10
        parts = []
        if full or hundred:
            parts.append(names[hundred])
            parts.append('trăm')
        if ten > 1:
            parts.append(names[ten])
            parts.append('mươi')
            if unit == 1:
                parts.append('mốt')
            elif unit == 5:
                parts.append('lăm')
            elif unit:
                parts.append(names[unit])
        elif ten == 1:
            parts.append('mười')
            if unit == 5:
                parts.append('lăm')
            elif unit:
                parts.append(names[unit])
        elif unit:
            if full or hundred:
                parts.append('lẻ')
            parts.append('năm' if unit == 5 and (full or hundred) else names[unit])
        return ' '.join(parts)

    units = ['', 'nghìn', 'triệu', 'tỷ']
    groups = []
    while number:
        groups.append(number % 1000)
        number //= 1000

    words = []
    for idx in range(len(groups) - 1, -1, -1):
        group = groups[idx]
        if group == 0:
            continue
        full = idx < len(groups) - 1
        words.append(read_three(group, full))
        if units[idx]:
            words.append(units[idx])
    return ' '.join(words)


def money_to_words(value: int) -> str:
    words = _digits_to_words(value).strip()
    if not words:
        words = 'không'
    return f'Bằng chữ: {words[:1].upper()}{words[1:]} đồng chẵn./.'


def safe_excel_filename(customer_name: str) -> str:
    safe_name = re.sub(r'[\\/:*?"<>|]', '-', customer_name or 'khach_hang')
    safe_name = re.sub(r'\s+', '_', safe_name).strip('_')[:40] or 'khach_hang'
    return f"bao_gia_{safe_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"


def build_quotation_excel(customer, products_qs, quote_number: str, custom_prices_map=None) -> bytes:
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    products = list(products_qs)
    extra_rows = max(0, len(products) - (PRODUCT_TEMPLATE_LAST_ROW - PRODUCT_START_ROW + 1))
    if extra_rows:
        _shift_merged_ranges(ws, TOTAL_ROW, extra_rows)
        ws.insert_rows(TOTAL_ROW, extra_rows)
        for row in range(TOTAL_ROW, TOTAL_ROW + extra_rows):
            _copy_row_style(ws, PRODUCT_TEMPLATE_LAST_ROW, row)

    total_row = TOTAL_ROW + extra_rows

    now = datetime.now()
    date_str = f"TP. Hồ Chí Minh, ngày {now.day:02d} tháng {now.month:02d} năm {now.year}"
    gia_label = PRICE_LABELS.get(customer.phan_loai, 'Giá DL+10%')
    if custom_prices_map:
        gia_label = 'Giá linh hoạt'

    ws['C3'] = date_str
    ws['A8'] = f"Khách hàng: {customer.ten_kh}"
    ws['A9'] = f"Địa chỉ: {customer.dia_chi or customer.tinh_tp or ''}"
    ws['A10'] = "Mã số thuế: "
    ws['A11'] = "Email: "

    for index, product in enumerate(products, start=1):
        row = PRODUCT_START_ROW + index - 1
        qty = 1
        
        if custom_prices_map and product.id in custom_prices_map:
            unit_price = Decimal(str(custom_prices_map[product.id]['price']))
        else:
            unit_price = product.get_price_for_type(customer.phan_loai) or Decimal('0')

        ws.cell(row, 1).value = index
        ws.cell(row, 2).value = f"{product.ma_vt} - {product.ten_hang or product.model_turbo or ''}"
        ws.cell(row, 3).value = product.dvt or 'Cái'
        ws.cell(row, 4).value = qty
        ws.cell(row, 5).value = int(unit_price)
        ws.cell(row, 6).value = float(VAT_RATE)
        ws.cell(row, 6).number_format = '0%'
        ws.cell(row, 7).value = f"=E{row}*D{row}*(1+F{row})"
        ws.cell(row, 7).number_format = '#,##0'

    for row in range(PRODUCT_START_ROW + len(products), total_row):
        ws.cell(row, 1).value = row - PRODUCT_START_ROW + 1
        for col in range(2, 8):
            ws.cell(row, col).value = None

    ws.cell(total_row, 1).value = 'TỔNG CỘNG'
    ws.cell(total_row, 7).value = f"=SUM(G13:G{total_row - 1})"
    ws.cell(total_row, 7).number_format = '#,##0'
    _configure_print_area(ws)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
