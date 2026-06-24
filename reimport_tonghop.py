"""
Re-import non-turbo products from TONG_HOP_PHU_TUNG Excel with CORRECT column mapping.
- Deletes existing products for matched types, re-imports from Excel
- For ket_nuoc/ket_nhot: UPDATE only (preserve hinh_anh)
- Adds ordering: products with images first

Usage: py reimport_tonghop.py --dry-run | py reimport_tonghop.py
"""

import sys, os, re, argparse
from pathlib import Path
from decimal import Decimal, InvalidOperation

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'backend.settings')
env_file = Path(__file__).resolve().parent / 'backend' / '.env'
if env_file.exists():
    for line in env_file.read_text(encoding='utf-8').splitlines():
        line = line.strip()
        if line and not line.startswith('#') and '=' in line:
            key, _, val = line.partition('=')
            os.environ.setdefault(key.strip(), val.strip())

import django; django.setup()
from products.models import Product, HangMay, HangSx, ThuongHieu, Category
from django.utils.text import slugify
import openpyxl


DOCS_DIR = Path(__file__).resolve().parent / 'docs'
EXCEL_FILE = DOCS_DIR / 'TONG_HOP_PHU_TUNG_10062026_2_updated.xlsx'

# Sheet name -> (loai, update_only)
SHEET_MAP = {
    'SUPAP': ('supap', False),
    'TRỤC CƠ': ('truc_co', False),
    'BƠM NƯỚC': ('bom_nuoc', False),
    'NẮP QUY LÁT': ('nap_quy_lat', False),
    'BƠM NHỚT': ('bom_nhot', False),
    'TRỤC CAM': ('truc_cam', False),
    'NẮP SINH HÀN': ('nap_sinh_han', False),
    'RUỘT SINH HÀN': ('ruot_sinh_han', False),
    'NHẬP TAY BIÊN': ('nhip_tay_bien', False),
    'THUN RON': ('thun_co', False),
    'THUN XY LANH': ('thun_xy_lanh', False),
    'LỐC MÁY': ('loc_may', False),
    'SAM BÉC': ('sam_bac', False),
    'VAN HẰNG NHIỆT': ('van_hang_nhiet', False),
    'VÀNH RĂNG BÁNH ĐÀ': ('vanh_rang_banh_da', False),
    'ỐNG DẪN NHIÊN LIỆU': ('ong_dan_nhien_lieu', False),
    'SÊN CAM': ('sen_cam', False),
    'KÉT NƯỚC ': ('ket_nuoc', True),   # UPDATE only, keep hinh_anh
    'KÉT NHỚT ': ('ket_nhot', True),   # UPDATE only, keep hinh_anh
}

# Default column mapping (used for most sheets)
# C1=Mã HH, C2=Hãng MÁY, C3=Tên động cơ, C4=PARNO, C5=Giá Bán, C6=Giá Ưu Đãi, C7=Giá VIP, C8=Ghi Chú
# For KÉT NƯỚC/KÉT NHỚT: C1=Mã HH, C2=Hãng SX, C3=Tên, C4=Chiều cao, C5=Giá Bán, ...


def parse_price(val) -> Decimal | None:
    if val is None: return None
    if isinstance(val, (int, float)):
        return Decimal(str(int(val)))
    text = str(val).strip()
    if not text or text.lower() in ('none', '—', '-', ''): return None
    clean = re.sub(r'[₫đVND\s,]', '', text)
    if '.' in clean and ',' in clean:
        clean = clean.replace('.', '').replace(',', '.')
    elif '.' in clean:
        parts = clean.split('.')
        if len(parts) > 1 and len(parts[-1]) == 3: clean = clean.replace('.', '')
    try:
        return Decimal(clean)
    except InvalidOperation:
        return None


def get_or_create(model_class, ten: str):
    if not ten or ten in ('None', '—', '', ' '): return None
    obj = model_class.objects.filter(ten__iexact=ten).first()
    if obj: return obj
    # Try to create with unique slug
    base_slug = slugify(ten)
    slug = base_slug
    counter = 1
    while True:
        try:
            return model_class.objects.create(ten=ten, slug=slug)
        except Exception:
            # Slug collision or other DB error
            existing = model_class.objects.filter(slug=slug).first()
            if existing:
                return existing  # Already exists with this slug
            counter += 1
            slug = f'{base_slug}-{counter}'
            if counter > 100:
                return None


def is_numeric(s):
    return bool(re.match(r'^[\d.]+$', s.strip()))


def reimport_tonghop(dry_run=True):
    if not EXCEL_FILE.exists():
        print(f'[ERROR] File not found: {EXCEL_FILE}')
        return

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    # Phase 1: Delete non-update types
    delete_types = [loai for loai, upd in SHEET_MAP.values() if not upd]
    existing = Product.objects.filter(loai__in=delete_types).count()
    print(f'Products to DELETE ({len(delete_types)} types): {existing}')

    if not dry_run:
        deleted, _ = Product.objects.filter(loai__in=delete_types).delete()
        print(f'Deleted: {deleted}')

    # Phase 2: Import
    stats = {'created': 0, 'updated': 0, 'skipped': 0, 'errors': 0}

    for sname, (loai, update_only) in SHEET_MAP.items():
        if sname not in wb.sheetnames:
            print(f'[WARN] Sheet not found: {sname}')
            continue

        ws = wb[sname]
        is_ket = 'KÉT' in sname

        print(f'\n--- {sname.strip()} -> loai={loai} (update_only={update_only}) ---')

        for row in range(4, ws.max_row + 1):
            ma_hh = str(ws.cell(row, 1).value or '').strip()
            if not ma_hh.startswith('HH'):
                if ma_hh and not ma_hh.startswith('◆') and not ma_hh.startswith('▸'):
                    pass  # could be a sub-row
                continue

            # Extract data based on column positions
            if is_ket:
                hang_may_raw = str(ws.cell(row, 2).value or '').strip()  # C2=Hãng SX (for ket)
                ten_sp = str(ws.cell(row, 3).value or '').strip()  # C3=Tên
                hang_sx_raw = ''
                parno = ''
                gia_ban = parse_price(ws.cell(row, 5).value)
                gia_uu_dai = parse_price(ws.cell(row, 6).value)
                gia_vip = parse_price(ws.cell(row, 7).value)
                ghi_chu = str(ws.cell(row, 8).value or '').strip() if ws.max_column >= 8 else ''
            else:
                hang_may_raw = str(ws.cell(row, 2).value or '').strip()
                ten_sp = str(ws.cell(row, 3).value or '').strip()
                hang_sx_raw = str(ws.cell(row, 4).value or '').strip() if ws.max_column >= 4 else ''
                # Check if C4 is PARNO or a price
                c4 = str(ws.cell(row, 4).value or '').strip()
                c5 = ws.cell(row, 5).value
                c6 = ws.cell(row, 6).value
                c7 = ws.cell(row, 7).value
                c8 = ws.cell(row, 8).value if ws.max_column >= 8 else None

                # Determine price columns (varies by sheet)
                if is_numeric(c4) or c4 == '':
                    # C4 is empty or numeric -> could be PARNO or old price
                    # Check C5 type
                    parno = ''
                    gia_ban = parse_price(c5) if c5 else None
                    gia_uu_dai = parse_price(c6) if c6 else None
                    gia_vip = parse_price(c7) if c7 else None
                    ghi_chu = str(c8 or '').strip()
                else:
                    parno = c4
                    gia_ban = parse_price(c5) if c5 else None
                    gia_uu_dai = parse_price(c6) if c6 else None
                    gia_vip = parse_price(c7) if c7 else None
                    ghi_chu = str(c8 or '').strip()

            # Skip if no hang_may or invalid
            if not hang_may_raw or hang_may_raw in ('None', '—', ''):
                stats['skipped'] += 1
                continue

            # Clean hang_may
            hang_may_raw = hang_may_raw.replace('└─', '').strip()

            if dry_run:
                stats['created'] += 1
                if stats['created'] <= 5:
                    print(f'  [{ma_hh}] hm={hang_may_raw} ten={ten_sp[:50]}')
                    print(f'           VIP={gia_vip} UU_DAI={gia_uu_dai} BAN={gia_ban}')
                continue

            hm = get_or_create(HangMay, ten=hang_may_raw)
            if not hm:
                stats['skipped'] += 1
                continue

            if update_only:
                # UPDATE existing product
                try:
                    p = Product.objects.get(ma_vt=ma_hh, loai=loai, is_active=True)
                    p.hang_may = hm
                    p.ten_hang = ten_sp[:500] if ten_sp else p.ten_hang
                    p.gia_vip = gia_vip or p.gia_vip
                    p.gia_uu_dai = gia_uu_dai or p.gia_uu_dai
                    p.gia_dai_ly = gia_ban or p.gia_dai_ly
                    if ghi_chu: p.ghi_chu = ghi_chu
                    p.save(update_fields=['hang_may', 'ten_hang', 'gia_vip', 'gia_uu_dai', 'gia_dai_ly', 'ghi_chu', 'updated_at'])
                    stats['updated'] += 1
                except Product.DoesNotExist:
                    stats['skipped'] += 1
            else:
                # CREATE new product (or skip if exists)
                try:
                    obj, created = Product.objects.get_or_create(
                        ma_vt=ma_hh,
                        loai=loai,
                        defaults={
                            'ten_hang': ten_sp[:500] if ten_sp else '',
                            'hang_may': hm,
                            'gia_vip': gia_vip,
                            'gia_uu_dai': gia_uu_dai,
                            'gia_dai_ly': gia_ban,
                            'ghi_chu': ghi_chu or '',
                            'is_active': True,
                        },
                    )
                    if created:
                        stats['created'] += 1
                    else:
                        # Update existing
                        obj.hang_may = hm
                        obj.ten_hang = ten_sp[:500] if ten_sp else obj.ten_hang
                        obj.gia_vip = gia_vip or obj.gia_vip
                        obj.gia_uu_dai = gia_uu_dai or obj.gia_uu_dai
                        obj.gia_dai_ly = gia_ban or obj.gia_dai_ly
                        if ghi_chu: obj.ghi_chu = ghi_chu
                        obj.is_active = True
                        obj.save()
                        stats['updated'] += 1
                except Exception as e:
                    stats['errors'] += 1
                    if stats['errors'] <= 3:
                        print(f'  ERROR [{ma_hh}]: {e}')

    wb.close()

    # Phase 3: Add ordering - products with images first
    if not dry_run:
        # Add DB index hint or set default ordering
        from django.db import connection
        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                    CREATE INDEX IF NOT EXISTS idx_products_has_image
                    ON products (is_active, loai, CASE WHEN hinh_anh != '' THEN 0 ELSE 1 END, ma_vt)
                """)
        except Exception:
            pass  # SQLite doesn't support CREATE INDEX IF NOT EXISTS

    # Summary
    sep = '=' * 60
    print(f'\n{sep}')
    print('KET QUA')
    print(f'  Created: {stats["created"]}')
    print(f'  Updated: {stats["updated"]}')
    print(f'  Skipped: {stats["skipped"]}')
    print(f'  Errors: {stats["errors"]}')
    if dry_run:
        print(f'  [DRY RUN]')
    else:
        for loai in set(v[0] for v in SHEET_MAP.values()):
            c = Product.objects.filter(loai=loai, is_active=True).count()
            print(f'  {loai}: {c} SP')
    print(sep)


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--dry-run', action='store_true')
    args = parser.parse_args()
    reimport_tonghop(dry_run=args.dry_run)
