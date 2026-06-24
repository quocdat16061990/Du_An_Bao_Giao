"""
Import ALL product data from BANG_GIA_BO_HOI_MOI_14.xlsx — the master data file.
Handles 13 data sheets with different column layouts.

Usage: py reimport_bohoi.py --dry-run | py reimport_bohoi.py
"""
import sys, os, re, argparse
from pathlib import Path
from decimal import Decimal, InvalidOperation

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'backend.settings')
env_file = Path(__file__).resolve().parent / 'backend' / '.env'
if env_file.exists():
    for line in env_file.read_text(encoding='utf-8').splitlines():
        line = line.strip()
        if line and not line.startswith('#') and '=' in line:
            key, _, val = line.partition('=')
            os.environ.setdefault(key.strip(), val.strip())

import django; django.setup()
from products.models import Product, HangMay, ThuongHieu, Category
from django.utils.text import slugify
import openpyxl

EXCEL = Path(__file__).resolve().parent / 'docs' / 'BANG_GIA_BO_HOI_MOI_14.xlsx'

# ── Sheet definitions: (sheet_name, loai, col_map, skip_rows_top) ──
# col_map: {db_field: excel_col_1indexed, ...}

SHEET_DEFS = [
    # RON BỘ: C1=STT, C2=Mã VT, C3=Tên hàng, C4=Hãng, C5=TH SX, C6=ĐVT, C7=Giá vốn, C8=VIP, C9=Ưu đãi, C10=Đại lý, C11=Gara
    {
        'sheet': '📦 RON BỘ', 'loai': 'ron_bo',
        'col_ma_vt': 2, 'col_hang_may': 4, 'col_ten': 3, 'col_th': 5,
        'col_gia_von': 7, 'col_gia_vip': 8, 'col_gia_uu_dai': 9, 'col_gia_dai_ly': 10, 'col_gia_gara': 11,
        'col_dvt': 6,
    },
    # RON MIẾNG: same structure
    {
        'sheet': '📄 RON MIẾNG', 'loai': 'ron_mieng',
        'col_ma_vt': 2, 'col_hang_may': 4, 'col_ten': 3, 'col_th': 5,
        'col_gia_von': 7, 'col_gia_vip': 8, 'col_gia_uu_dai': 9, 'col_gia_dai_ly': 10, 'col_gia_gara': 11,
        'col_dvt': 6,
    },
    # MIỂNG TNC: C1=Mã HH, C2=Hãng, C3=Mã ĐC, C4=ĐK, C5=Cos, C6=Loại, C7=Thương hiệu, C8=ĐVT
    {
        'sheet': '📋 MIỂNG TNC', 'loai': 'mieng_bac',
        'col_ma_vt': 1, 'col_hang_may': 2, 'col_ten': 3, 'col_th': 7,
        'col_ma_dong_co': 3, 'col_dvt': 8,
        'extra_cols': {4: 'ĐK', 5: 'Cos', 6: 'Loại'},
    },
    # BẠC THAU: C1=Mã HH, C2=Tên vật tư, C3=ĐVT, C4=Vốn, C5=VIP, C6=Ưu đãi, C7=Đại lý
    {
        'sheet': '🔧 BẠC THAU', 'loai': 'can_thau',
        'col_ma_vt': 1, 'col_ten': 2, 'col_dvt': 3,
        'col_gia_von': 4, 'col_gia_vip': 5, 'col_gia_uu_dai': 6, 'col_gia_dai_ly': 7,
    },
    # CĂN DỌC: same as BẠC THAU
    {
        'sheet': '📏 CĂN DỌC', 'loai': 'can_doc',
        'col_ma_vt': 1, 'col_ten': 2, 'col_dvt': 3,
        'col_gia_von': 4, 'col_gia_vip': 5, 'col_gia_uu_dai': 6, 'col_gia_dai_ly': 7,
    },
    # PISTON: C1=Mã HH, C2=Hãng, C3=Mã ĐC, C4=ĐK, C5=Ký hiệu, C6=Buồng nổ, C7=Ắc
    {
        'sheet': '🔩 PISTON', 'loai': 'piston',
        'col_ma_vt': 1, 'col_hang_may': 2, 'col_ma_dong_co': 3, 'col_ten': 3,
        'extra_cols': {4: 'ĐK', 5: 'Ký hiệu', 6: 'Buồng nổ', 7: 'Ắc', 8: 'Ắc đỉnh'},
    },
    # SÉC MĂNG: C1=Mã HH, C2=Hãng, C3=Mã ĐC, C4=ĐK, C5=Kiếng?, C6=Loại mạ, C7=Ring, C8=Thương hiệu
    {
        'sheet': '⚙️ SÉC MĂNG', 'loai': 'sec_mang',
        'col_ma_vt': 1, 'col_hang_may': 2, 'col_ma_dong_co': 3, 'col_ten': 3, 'col_th': 8,
        'extra_cols': {4: 'ĐK', 5: 'Kiếng?', 6: 'Loại mạ', 7: 'Ring'},
    },
    # XY LANH: C1=Mã HH, C2=Hãng, C3=Mã ĐC, C4=Số máy, C5=Kiểu, C6=Xoáy, C7=Doa, C8=Thông số
    {
        'sheet': '🔧 XY LANH', 'loai': 'xy_lanh',
        'col_ma_vt': 1, 'col_hang_may': 2, 'col_ma_dong_co': 3, 'col_ten': 3,
        'extra_cols': {4: 'Số máy', 5: 'Kiểu', 6: 'Xoáy', 7: 'Doa', 8: 'Thông số'},
    },
    # XY LANH CŨ: C1=Mã HH, C2=Hãng, C3=Mã ĐC, C4=Số máy, C5=Kiểu, C6=Xoáy, C7=Doa
    {
        'sheet': '🗄️ XY LANH CŨ', 'loai': 'xy_lanh_cu',
        'col_ma_vt': 1, 'col_hang_may': 2, 'col_ma_dong_co': 3, 'col_ten': 3,
        'extra_cols': {4: 'Số máy', 5: 'Kiểu', 6: 'Xoáy', 7: 'Doa', 8: 'Thông số'},
    },
    # RON CẠT TE: C1=Mã HH, C2=Hãng, C3=Mã ĐC, C4=ĐK, C5=Thương hiệu, C6=ĐVT, C7=Vốn, C8=VIP
    {
        'sheet': '🛢️ RON CẠT TE', 'loai': 'ron_cat_te',
        'col_ma_vt': 1, 'col_hang_may': 2, 'col_ma_dong_co': 3, 'col_ten': 3, 'col_th': 5,
        'col_dvt': 6, 'col_gia_von': 7, 'col_gia_vip': 8, 'col_gia_uu_dai': 9, 'col_gia_dai_ly': 10,
        'extra_cols': {4: 'ĐK'},
    },
    # THUN CÒ
    {
        'sheet': '〰️ THUN CÒ', 'loai': 'thun_co',
        'col_ma_vt': 1, 'col_hang_may': 2, 'col_ma_dong_co': 3, 'col_ten': 3, 'col_th': 5,
        'col_dvt': 6, 'col_gia_von': 7, 'col_gia_vip': 8, 'col_gia_uu_dai': 9, 'col_gia_dai_ly': 10,
        'extra_cols': {4: 'ĐK'},
    },
    # PHỚT ĐẦU
    {
        'sheet': '⭕ PHỚT ĐẦU TRỤC CƠ', 'loai': 'phot_dau',
        'col_ma_vt': 1, 'col_hang_may': 2, 'col_ma_dong_co': 3, 'col_ten': 3, 'col_th': 6,
        'col_dvt': 7, 'col_gia_von': 8, 'col_gia_vip': 9, 'col_gia_uu_dai': 10, 'col_gia_dai_ly': 11,
        'extra_cols': {4: 'ĐK', 5: 'Kích thước'},
    },
    # PHỚT ĐUÔI
    {
        'sheet': '🔴 PHỚT ĐUÔI TRỤC CƠ', 'loai': 'phot_duoi',
        'col_ma_vt': 1, 'col_hang_may': 2, 'col_ma_dong_co': 3, 'col_ten': 3, 'col_th': 6,
        'col_dvt': 7, 'col_gia_von': 8, 'col_gia_vip': 9, 'col_gia_uu_dai': 10, 'col_gia_dai_ly': 11,
        'extra_cols': {4: 'ĐK', 5: 'Kích thước'},
    },
]


def parse_price(val):
    if val is None: return None
    if isinstance(val, (int, float)):
        return Decimal(str(int(val)))
    text = str(val).strip()
    if not text or text.lower() in ('none', '—', '-'): return None
    clean = re.sub(r'[₫đVND\s,A-Za-z]', '', text)
    if not clean: return None
    if '.' in clean:
        parts = clean.split('.')
        if len(parts) > 1 and len(parts[-1]) == 3:
            clean = clean.replace('.', '')
    try: return Decimal(clean)
    except InvalidOperation: return None


def get_or_create(model_class, ten: str):
    if not ten or ten in ('None', '—', ''): return None
    obj = model_class.objects.filter(ten__iexact=ten).first()
    if obj: return obj
    base_slug = slugify(ten)
    slug = base_slug
    for _ in range(10):
        try: return model_class.objects.create(ten=ten, slug=slug)
        except Exception:
            existing = model_class.objects.filter(slug=slug).first()
            if existing: return existing
            slug = f'{base_slug}-{_}'
    return None


def get_cell(ws, row, col, default=''):
    """Get cell value safely."""
    if col is None or col > ws.max_column: return default
    v = ws.cell(row, col).value
    return str(v).strip() if v is not None else default


def import_bohoi(dry_run=True):
    if not EXCEL.exists():
        print(f'[ERROR] File not found: {EXCEL}'); return

    wb = openpyxl.load_workbook(EXCEL, data_only=True)

    # Phase 1: Delete existing products for these types
    all_loai = list(set(d['loai'] for d in SHEET_DEFS))
    existing = Product.objects.filter(loai__in=all_loai).count()
    print(f'Products to DELETE ({len(all_loai)} types): {existing}')

    if not dry_run:
        deleted, _ = Product.objects.filter(loai__in=all_loai).delete()
        print(f'Deleted: {deleted}')

    # Also create known ThuongHieu from CHÚ GIẢI
    KNOWN_TH = ['NPR', 'RIK', 'Rikken', 'TP', 'IKAZU', 'IZUMI', 'HENWEIT', 'NSP', 'MAHLE', 'PROXMANN']
    for th_name in KNOWN_TH:
        get_or_create(ThuongHieu, ten=th_name)

    stats = {'created': 0, 'skipped': 0, 'errors': 0}

    for sdef in SHEET_DEFS:
        sname, loai = sdef['sheet'], sdef['loai']
        if sname not in wb.sheetnames:
            print(f'[WARN] Sheet not found: {sname}'); continue

        ws = wb[sname]
        print(f'Processing: {sname} -> {loai} ({ws.max_row} rows)')

        # Get or create category
        cat = get_or_create(Category, ten=sname.strip().split(' ')[-1] if ' ' in sname else sname)

        for row in range(2, ws.max_row + 1):
            ma_vt = get_cell(ws, row, sdef.get('col_ma_vt', 1))

            # Skip non-HH rows
            if not ma_vt.startswith('HH'):
                continue

            hang_may_raw = get_cell(ws, row, sdef.get('col_hang_may'))
            ten_sp = get_cell(ws, row, sdef.get('col_ten'))

            # Skip if no meaningful data
            if not ten_sp and not hang_may_raw:
                stats['skipped'] += 1; continue

            # Price columns (optional)
            gia_von = parse_price(ws.cell(row, sdef['col_gia_von']).value) if sdef.get('col_gia_von') else None
            gia_vip = parse_price(ws.cell(row, sdef['col_gia_vip']).value) if sdef.get('col_gia_vip') else None
            gia_uu_dai = parse_price(ws.cell(row, sdef['col_gia_uu_dai']).value) if sdef.get('col_gia_uu_dai') else None
            gia_dai_ly = parse_price(ws.cell(row, sdef['col_gia_dai_ly']).value) if sdef.get('col_gia_dai_ly') else None
            gia_gara = parse_price(ws.cell(row, sdef['col_gia_gara']).value) if sdef.get('col_gia_gara') else None

            ma_dong_co = get_cell(ws, row, sdef.get('col_ma_dong_co'))
            dvt = get_cell(ws, row, sdef.get('col_dvt'))

            # Extra attributes
            attrs = {}
            for col, label in sdef.get('extra_cols', {}).items():
                v = get_cell(ws, row, col)
                if v: attrs[label] = v

            if dry_run:
                stats['created'] += 1
                if stats['created'] <= 3:
                    print(f'  [{ma_vt}] hm={hang_may_raw} ten={ten_sp[:50]}')
                continue

            # Create product
            hm = get_or_create(HangMay, ten=hang_may_raw) if hang_may_raw else None
            th_raw = get_cell(ws, row, sdef.get('col_th'))
            th = get_or_create(ThuongHieu, ten=th_raw) if th_raw else None

            try:
                obj, created = Product.objects.get_or_create(
                    ma_vt=ma_vt, loai=loai,
                    defaults={
                        'ten_hang': ten_sp[:500] if ten_sp else '',
                        'hang_may': hm,
                        'thuong_hieu': th,
                        'category': cat,
                        'ma_dong_co': ma_dong_co or '',
                        'dvt': dvt or 'Cái',
                        'gia_von': gia_von,
                        'gia_vip': gia_vip,
                        'gia_uu_dai': gia_uu_dai,
                        'gia_dai_ly': gia_dai_ly,
                        'gia_gara': gia_gara,
                        'attributes': attrs,
                        'is_active': True,
                    }
                )
                if created:
                    stats['created'] += 1
                else:
                    # Update existing
                    if hm: obj.hang_may = hm
                    if th: obj.thuong_hieu = th
                    if ten_sp: obj.ten_hang = ten_sp[:500]
                    if gia_vip: obj.gia_vip = gia_vip
                    if gia_uu_dai: obj.gia_uu_dai = gia_uu_dai
                    if gia_dai_ly: obj.gia_dai_ly = gia_dai_ly
                    obj.is_active = True
                    obj.attributes = attrs
                    obj.save()
            except Exception as e:
                stats['errors'] += 1
                if stats['errors'] <= 3:
                    print(f'  ERROR [{ma_vt}]: {e}')

    wb.close()

    # Summary
    sep = '=' * 60
    print(f'\n{sep}')
    print('KET QUA')
    print(f'  Created: {stats["created"]}')
    print(f'  Skipped: {stats["skipped"]}')
    print(f'  Errors: {stats["errors"]}')
    if dry_run:
        print('  [DRY RUN]')
    else:
        for loai in all_loai:
            c = Product.objects.filter(loai=loai, is_active=True).count()
            if c > 0:
                print(f'  {loai}: {c} SP')
    print(sep)


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--dry-run', action='store_true')
    args = parser.parse_args()
    import_bohoi(dry_run=args.dry_run)
