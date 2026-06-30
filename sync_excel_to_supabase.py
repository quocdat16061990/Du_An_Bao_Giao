import sys
import os
import re
import argparse
import openpyxl
from pathlib import Path
from decimal import Decimal

# Configure stdout to handle Vietnamese characters correctly in Windows PowerShell
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

project_root = r"D:\Du_An_Bao_Giao"
if project_root not in sys.path:
    sys.path.insert(0, project_root)

# Set DJANGO_SETTINGS_MODULE and setup Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'backend.settings')

# Load backend/.env manually to read database configurations
env_file = Path(project_root) / 'backend' / '.env'
if env_file.exists():
    for line in env_file.read_text(encoding='utf-8').splitlines():
        line = line.strip()
        if line and not line.startswith('#') and '=' in line:
            key, _, val = line.partition('=')
            os.environ[key.strip()] = val.strip()

import django
django.setup()

from django.db import transaction
from django.utils.text import slugify
from products.models import Product, HangMay, ThuongHieu, Category

# High surrogate emojis regex
try:
    high_surrogates = re.compile(r'[\U00010000-\U0010ffff]', flags=re.UNICODE)
except re.error:
    high_surrogates = re.compile(r'[\uD800-\uDBFF][\uDC00-\uDFFF]')

other_symbols = re.compile(r'[\u2600-\u27BF\u2300-\u23FF\u2B50\u2190-\u21FF\u2B00-\u2BFF\u2900-\u297F\u25C6]', flags=re.UNICODE)

def remove_emoji(text):
    if not text:
        return ""
    text = str(text)
    text = high_surrogates.sub('', text)
    text = other_symbols.sub('', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

def parse_price(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return Decimal(str(int(val)))
    text = str(val).strip()
    if not text or text.lower() in ('none', '—', '-', ''):
        return None
    clean = re.sub(r'[₫đVND\s,A-Za-z]', '', text)
    if not clean:
        return None
    if '.' in clean:
        parts = clean.split('.')
        if len(parts) > 1 and len(parts[-1]) == 3:
            clean = clean.replace('.', '')
    try:
        return Decimal(clean)
    except:
        return None

def get_cell_val(ws, row, col):
    if col is None or col > ws.max_column:
        return ""
    v = ws.cell(row, col).value
    return str(v).strip() if v is not None else ""

def get_or_create_cached_lookup(model_class, ten, cache, dry_run=False):
    """Safely get or create a lookup item using slug-based caching to prevent slug unique key violations."""
    if not ten or ten in ('None', '—', ''):
        return None
    ten = remove_emoji(ten)
    if not ten:
        return None
    
    slug = slugify(ten) or 'mac-dinh'
    
    if slug in cache:
        return cache[slug]
        
    # Check DB by slug (most robust way to prevent UniqueViolation)
    obj = model_class.objects.filter(slug=slug).first()
    if obj:
        cache[slug] = obj
        return obj
        
    if dry_run:
        mock = model_class(ten=ten, slug=slug)
        cache[slug] = mock
        return mock
        
    # Create in DB and cache it
    for i in range(100):
        try:
            obj = model_class.objects.create(ten=ten, slug=slug)
            cache[slug] = obj
            return obj
        except Exception:
            # Handle potential race conditions or concurrent collisions by changing slug suffix
            slug = f"{slugify(ten)}-{i}"
            existing = model_class.objects.filter(slug=slug).first()
            if existing:
                cache[slug] = existing
                return existing
    return None

def draw_progress_bar(iteration, total, prefix='', suffix='', length=30, fill='█'):
    """Draws a premium dynamic terminal progress bar."""
    percent = f"{100 * (iteration / float(total)):.1f}%"
    filled_length = int(length * iteration // total)
    bar = fill * filled_length + '-' * (length - filled_length)
    sys.stdout.write(f"\r{prefix} |{bar}| {percent} {suffix}")
    sys.stdout.flush()
    if iteration == total:
        sys.stdout.write('\n')
        sys.stdout.flush()

SHEET_DEFS = [
    {
        'sheet': 'PISTON', 'loai': 'piston', 'col_ma_vt': 1, 'col_hang_may': 2, 'col_ten': 3, 'col_th': 11, 'col_dvt': 12,
        'col_gia_vip': 14, 'col_gia_uu_dai': 15, 'col_gia_dai_ly': 16, 'col_gia_gara': 17, 'col_parno': 19, 'col_ghi_chu': 20,
        'extra_cols': {4: 'ĐK', 5: 'Ký hiệu', 6: 'Buồng nổ', 7: 'Ắc', 8: 'Ắc đỉnh', 9: 'Tổng dài', 10: 'Ring 1-5'}
    },
    {
        'sheet': 'SÉC MĂNG', 'loai': 'sec_mang', 'col_ma_vt': 1, 'col_hang_may': 2, 'col_ten': 3, 'col_th': 8, 'col_dvt': 9,
        'col_gia_vip': 11, 'col_gia_uu_dai': 12, 'col_gia_dai_ly': 13, 'col_gia_gara': 14, 'col_parno': 16, 'col_ghi_chu': 17,
        'extra_cols': {4: 'ĐK', 5: 'Xy lanh kiếng?', 6: 'Loại', 7: 'Ring'}
    },
    {
        'sheet': 'XY LANH', 'loai': 'xy_lanh', 'col_ma_vt': 1, 'col_hang_may': 2, 'col_ten': 3, 'col_th': 8, 'col_dvt': 9,
        'col_gia_vip': 10, 'col_gia_uu_dai': 11, 'col_gia_dai_ly': 12, 'col_gia_gara': 13, 'col_parno': 7,
        'extra_cols': {4: 'Số máy', 5: 'LOẠI', 17: 'Gợi ý Mã HH'}
    },
    {
        'sheet': 'THUN CÒ', 'loai': 'thun_co', 'col_ma_vt': 1, 'col_hang_may': 2, 'col_ten': 15, 'col_th': 5, 'col_dvt': 6,
        'col_gia_vip': 8, 'col_gia_uu_dai': 9, 'col_gia_dai_ly': 10, 'col_gia_gara': 11, 'col_parno': 13, 'col_ghi_chu': 14,
        'extra_cols': {4: 'ĐK'}
    },
    {
        'sheet': 'PHỚT ĐẦU TRỤC CƠ', 'loai': 'phot_dau', 'col_ma_vt': 1, 'col_hang_may': 2, 'col_ten': 16, 'col_th': 6, 'col_dvt': 7,
        'col_gia_vip': 9, 'col_gia_uu_dai': 10, 'col_gia_dai_ly': 11, 'col_gia_gara': 12, 'col_parno': 14, 'col_ghi_chu': 15,
        'extra_cols': {4: 'ĐK', 5: 'Kích thước'}
    },
    {
        'sheet': 'PHỚT ĐUÔI TRỤC CƠ', 'loai': 'phot_duoi', 'col_ma_vt': 1, 'col_hang_may': 2, 'col_ten': 16, 'col_th': 6, 'col_dvt': 7,
        'col_gia_vip': 9, 'col_gia_uu_dai': 10, 'col_gia_dai_ly': 11, 'col_gia_gara': 12, 'col_parno': 14, 'col_ghi_chu': 15,
        'extra_cols': {4: 'ĐK', 5: 'Kích thước'}
    },
    {
        'sheet': 'RON BỘ', 'loai': 'ron_bo', 'col_ma_vt': 2, 'col_hang_may': 5, 'col_ten': 3, 'col_th': 6, 'col_dvt': 7,
        'col_gia_vip': 9, 'col_gia_uu_dai': 10, 'col_gia_dai_ly': 11, 'col_gia_gara': 12, 'col_ghi_chu': 14,
        'extra_cols': {4: 'ĐK'}
    },
    {
        'sheet': 'RON MIẾNG', 'loai': 'ron_mieng', 'col_ma_vt': 2, 'col_hang_may': 5, 'col_ten': 3, 'col_th': 6, 'col_dvt': 7,
        'col_gia_vip': 10, 'col_gia_uu_dai': 11, 'col_gia_dai_ly': 12, 'col_gia_gara': 13, 'col_ghi_chu': 15,
        'extra_cols': {4: 'ĐK'}
    },
    {
        'sheet': 'MIỂNG TNC', 'loai': 'mieng_bac', 'col_ma_vt': 1, 'col_hang_may': 2, 'col_ten': 16, 'col_th': 7, 'col_dvt': 8,
        'col_gia_vip': 10, 'col_gia_uu_dai': 11, 'col_gia_dai_ly': 12, 'col_gia_gara': 13, 'col_parno': 15, 'col_ghi_chu': 9,
        'extra_cols': {4: 'ĐK', 5: 'Cos', 6: 'Loại'}
    },
    {
        'sheet': 'BẠC THAU', 'loai': 'can_thau', 'col_ma_vt': 1, 'col_ten': 2, 'col_dvt': 4,
        'col_gia_vip': 6, 'col_gia_uu_dai': 7, 'col_gia_dai_ly': 8, 'col_gia_gara': 9, 'col_parno': 11, 'col_ghi_chu': 12,
        'extra_cols': {3: 'ĐK'}
    },
    {
        'sheet': 'CĂN DỌC', 'loai': 'can_doc', 'col_ma_vt': 1, 'col_ten': 2, 'col_dvt': 4,
        'col_gia_vip': 6, 'col_gia_uu_dai': 7, 'col_gia_dai_ly': 8, 'col_gia_gara': 9, 'col_parno': 11, 'col_ghi_chu': 12,
        'extra_cols': {3: 'ĐK'}
    }
]

def main():
    parser = argparse.ArgumentParser(description="Synchronize Excel products data to Supabase database.")
    parser.add_argument('--real-run', action='store_true', help='Perform actual database write, otherwise defaults to dry-run')
    args = parser.parse_args()
    
    dry_run = not args.real_run
    
    if dry_run:
        print("="*60)
        print(" RUNNING IN DRY-RUN MODE — NO DATABASE CHANGES WILL BE APPLIED ")
        print("="*60)
    else:
        print("!"*60)
        print(" WARNING: REAL RUN MODE — DATABASE CHANGES WILL BE PERSISTED ")
        print("!"*60)

    excel_path = r"D:\Du_An_Bao_Giao\docs\BAO GIA BO HOI - SEC MANG (DA DO GIA PISTON).xlsx"
    if not os.path.exists(excel_path):
        print(f"[ERROR] Excel file not found at: {excel_path}")
        return

    # 1. Preload all products in DB for O(1) matching
    print("Preloading database records into RAM...")
    db_prods = Product.objects.filter(is_active=True)
    db_dict = {}
    for p in db_prods:
        db_dict[(p.ma_vt, p.loai)] = p
    print(f"Preloaded {len(db_dict)} active product variants.")

    # 2. Preload lookup tables for cache - KEYED BY SLUGIFY(NAME) TO PREVENT COLLISONS
    print("Caching lookup tables (HangMay, ThuongHieu, Category)...")
    hang_may_cache = {}
    for hm in HangMay.objects.all():
        s = slugify(hm.ten)
        if s: hang_may_cache[s] = hm
        
    thuong_hieu_cache = {}
    for th in ThuongHieu.objects.all():
        s = slugify(th.ten)
        if s: thuong_hieu_cache[s] = th
        
    category_cache = {}
    for cat in Category.objects.all():
        s = slugify(cat.ten)
        if s: category_cache[s] = cat
    
    fallback_brand = get_or_create_cached_lookup(HangMay, "KHÁC", hang_may_cache, dry_run=dry_run)

    print("Loading Excel file (openpyxl)...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    stats = {
        'total_excel': 0,
        'created': 0,
        'updated_prices': 0,
        'updated_fields': 0,
        'updated_attributes': 0,
        'unchanged': 0,
        'errors': 0
    }
    
    report_lines = []
    report_lines.append("# BÁO CÁO KẾT QUẢ ĐỒNG BỘ DỮ LIỆU EXCEL LÊN SUPABASE")
    report_lines.append(f"\n- **Chế độ thực thi:** `{'DRY-RUN (Thử nghiệm)' if dry_run else 'REAL-RUN (Ghi thực tế)'}`")
    report_lines.append(f"- **Tệp nguồn:** `{Path(excel_path).name}`\n")
    
    sheet_reports = []

    for sdef in SHEET_DEFS:
        sheet_clean_name = sdef['sheet']
        loai = sdef['loai']
        
        # Find sheet
        actual_name = None
        for name in wb.sheetnames:
            if sheet_clean_name in remove_emoji(name).upper():
                actual_name = name
                break
                
        if not actual_name:
            print(f"[WARN] Skip sheet {sheet_clean_name}: sheet not found in Excel")
            continue
            
        print(f"\nProcessing sheet: {actual_name}...")
        ws = wb[actual_name]
        total_rows = ws.max_row
        
        # Find header row
        header_row_idx = 0
        max_non_empty = 0
        for ri in range(min(5, total_rows)):
            row_vals = [ws.cell(row=ri+1, column=c).value for c in range(1, ws.max_column + 1)]
            cnt = sum(1 for v in row_vals if v and str(v).strip())
            if cnt > max_non_empty:
                max_non_empty = cnt
                header_row_idx = ri
                
        # Group category matching
        cat_name = sheet_clean_name
        cat = get_or_create_cached_lookup(Category, cat_name, category_cache, dry_run=dry_run)
        
        sheet_stats = {'total': 0, 'created': 0, 'updated': 0, 'unchanged': 0, 'errors': 0}
        details_list = []
        
        # Lists for bulk database updates to minimize remote query count
        to_create_list = []
        to_update_list = []
        
        start_row = header_row_idx + 2
        total_data_rows = total_rows - start_row + 1
        
        for r_idx in range(start_row, total_rows + 1):
            if total_data_rows > 0:
                draw_progress_bar(r_idx - start_row, total_data_rows, prefix=f"  {sheet_clean_name:<16}", suffix=f"Row {r_idx}/{total_rows}")
                
            ma_vt = get_cell_val(ws, r_idx, sdef['col_ma_vt'])
            if not ma_vt or not ma_vt.startswith('HH'):
                continue
                
            sheet_stats['total'] += 1
            stats['total_excel'] += 1
            
            # Read variables
            hang_may_raw = get_cell_val(ws, r_idx, sdef.get('col_hang_may'))
            ten_sp_raw = get_cell_val(ws, r_idx, sdef.get('col_ten'))
            th_raw = get_cell_val(ws, r_idx, sdef.get('col_th'))
            dvt_raw = get_cell_val(ws, r_idx, sdef.get('col_dvt')) or 'Cái'
            parno_raw = get_cell_val(ws, r_idx, sdef.get('col_parno'))
            ghi_chu_raw = get_cell_val(ws, r_idx, sdef.get('col_ghi_chu'))
            
            # Prices
            gia_vip = parse_price(ws.cell(r_idx, sdef['col_gia_vip']).value) if sdef.get('col_gia_vip') else None
            gia_uu_dai = parse_price(ws.cell(r_idx, sdef['col_gia_uu_dai']).value) if sdef.get('col_gia_uu_dai') else None
            gia_dai_ly = parse_price(ws.cell(r_idx, sdef['col_gia_dai_ly']).value) if sdef.get('col_gia_dai_ly') else None
            gia_gara = parse_price(ws.cell(r_idx, sdef['col_gia_gara']).value) if sdef.get('col_gia_gara') else None
            
            # Extra attributes
            extra_attrs = {}
            for col_idx, key_name in sdef.get('extra_cols', {}).items():
                val = get_cell_val(ws, r_idx, col_idx)
                if val:
                    extra_attrs[key_name] = val
            
            # Check exist in DB cache
            p_db = db_dict.get((ma_vt, loai))
            
            try:
                if not p_db:
                    # CREATE PRODUCT IN-MEMORY
                    hm = get_or_create_cached_lookup(HangMay, hang_may_raw, hang_may_cache, dry_run=dry_run) if hang_may_raw else None
                    th = get_or_create_cached_lookup(ThuongHieu, th_raw, thuong_hieu_cache, dry_run=dry_run) if th_raw else None
                    
                    if not hm:
                        hm = fallback_brand
                        
                    new_prod = Product(
                        ma_vt=ma_vt,
                        loai=loai,
                        ten_hang=ten_sp_raw[:500] if ten_sp_raw else ma_vt,
                        hang_may=hm,
                        thuong_hieu=th,
                        category=cat,
                        dvt=dvt_raw[:50],
                        parno=parno_raw[:300] if parno_raw else '',
                        ghi_chu=ghi_chu_raw or '',
                        gia_vip=gia_vip,
                        gia_uu_dai=gia_uu_dai,
                        gia_dai_ly=gia_dai_ly,
                        gia_gara=gia_gara,
                        attributes=extra_attrs,
                        is_active=True
                    )
                    to_create_list.append(new_prod)
                    sheet_stats['created'] += 1
                    stats['created'] += 1
                    details_list.append(f"- `[HH] CREATE`: `[{ma_vt}]` {ten_sp_raw[:40]} | Giá VIP: `{gia_vip}` | Hãng: `{hang_may_raw}`")
                
                else:
                    # UPDATE/FILL PRODUCT IN-MEMORY
                    is_changed = False
                    change_reasons = []
                    
                    # 1. Update prices if empty or if different (Excel is priority)
                    if gia_vip is not None and p_db.gia_vip != gia_vip:
                        p_db.gia_vip = gia_vip
                        is_changed = True
                        change_reasons.append("Cập nhật Giá VIP")
                    if gia_uu_dai is not None and p_db.gia_uu_dai != gia_uu_dai:
                        p_db.gia_uu_dai = gia_uu_dai
                        is_changed = True
                        change_reasons.append("Cập nhật Giá Ưu đãi")
                    if gia_dai_ly is not None and p_db.gia_dai_ly != gia_dai_ly:
                        p_db.gia_dai_ly = gia_dai_ly
                        is_changed = True
                        change_reasons.append("Cập nhật Giá Đại lý")
                    if gia_gara is not None and p_db.gia_gara != gia_gara:
                        p_db.gia_gara = gia_gara
                        is_changed = True
                        change_reasons.append("Cập nhật Giá Gara")
                        
                    # 2. Fill basic fields if empty in DB
                    if not p_db.dvt and dvt_raw:
                        p_db.dvt = dvt_raw[:50]
                        is_changed = True
                        change_reasons.append("Bổ sung ĐVT")
                    if not p_db.parno and parno_raw:
                        p_db.parno = parno_raw[:300]
                        is_changed = True
                        change_reasons.append("Bổ sung Part Number")
                    if not p_db.ghi_chu and ghi_chu_raw:
                        p_db.ghi_chu = ghi_chu_raw
                        is_changed = True
                        change_reasons.append("Bổ sung Ghi chú")
                    if not p_db.ten_hang and ten_sp_raw:
                        p_db.ten_hang = ten_sp_raw[:500]
                        is_changed = True
                        change_reasons.append("Bổ sung Tên hàng")
                        
                    # 3. Fill FK associations if empty
                    if not p_db.hang_may and hang_may_raw:
                        hm = get_or_create_cached_lookup(HangMay, hang_may_raw, hang_may_cache, dry_run=dry_run)
                        p_db.hang_may = hm
                        is_changed = True
                        change_reasons.append("Bổ sung Hãng máy")
                    if not p_db.thuong_hieu and th_raw:
                        th = get_or_create_cached_lookup(ThuongHieu, th_raw, thuong_hieu_cache, dry_run=dry_run)
                        p_db.thuong_hieu = th
                        is_changed = True
                        change_reasons.append("Bổ sung Thương hiệu")
                    if not p_db.category and cat:
                        p_db.category = cat
                        is_changed = True
                        change_reasons.append("Bổ sung Category")
                        
                    # 4. Merge technical attributes in JSON
                    if extra_attrs:
                        db_attrs = p_db.attributes or {}
                        merged_attrs = {**db_attrs}
                        attrs_changed = False
                        for ak, av in extra_attrs.items():
                            if db_attrs.get(ak) != av:
                                merged_attrs[ak] = av
                                attrs_changed = True
                        if attrs_changed:
                            p_db.attributes = merged_attrs
                            is_changed = True
                            change_reasons.append("Cập nhật/Bổ sung Thuộc tính")
                            
                    if is_changed:
                        to_update_list.append(p_db)
                        sheet_stats['updated'] += 1
                        stats['updated_prices'] += 1
                        details_list.append(f"- `[HH] UPDATE`: `[{ma_vt}]` {p_db.display_name()[:40]} -> {', '.join(change_reasons)}")
                    else:
                        sheet_stats['unchanged'] += 1
                        stats['unchanged'] += 1
                        
            except Exception as e:
                sheet_stats['errors'] += 1
                stats['errors'] += 1
                details_list.append(f"- `[ERROR]`: `[{ma_vt}]` Lỗi: {e}")
                
        # Draw final 100% bar
        if total_data_rows > 0:
            draw_progress_bar(total_data_rows, total_data_rows, prefix=f"  {sheet_clean_name:<16}", suffix=f"Row {total_rows}/{total_rows}")

        # 5. EXECUTE DATABASE WRITES USING BULK OPERATIONS (HUGE LATENCY REDUCTION)
        if not dry_run:
            if to_create_list:
                Product.objects.bulk_create(to_create_list, batch_size=500)
            if to_update_list:
                Product.objects.bulk_update(to_update_list, fields=[
                    'gia_vip', 'gia_uu_dai', 'gia_dai_ly', 'gia_gara', 
                    'dvt', 'parno', 'ghi_chu', 'ten_hang', 
                    'hang_may', 'thuong_hieu', 'category', 'attributes'
                ], batch_size=500)

        sheet_reports.append(f"\n### Sheet: `{sheet_clean_name}` (Loại: `{loai}`)")
        sheet_reports.append(f"- **Tổng số dòng Excel quét:** `{sheet_stats['total']}`")
        sheet_reports.append(f"- **Thêm mới (INSERT):** `{sheet_stats['created']}`")
        sheet_reports.append(f"- **Cập nhật/Bổ sung (UPDATE):** `{sheet_stats['updated']}`")
        sheet_reports.append(f"- **Không đổi:** `{sheet_stats['unchanged']}`")
        sheet_reports.append(f"- **Số lượng dòng lỗi:** `{sheet_stats['errors']}`\n")
        
        if details_list:
            sheet_reports.append("#### Chi tiết thay đổi (Tối đa 20 dòng):")
            sheet_reports.extend(details_list[:20])
            if len(details_list) > 20:
                sheet_reports.append(f"- *...và còn {len(details_list) - 20} dòng thay đổi khác.*")

    wb.close()
    
    # Write Overall Summary
    summary_sect = []
    summary_sect.append("## BẢNG THỐNG KÊ TỔNG QUAN")
    summary_sect.append(f"- **Tổng số sản phẩm quét từ Excel:** `{stats['total_excel']}`")
    summary_sect.append(f"- **Số sản phẩm tạo mới (INSERT):** `{stats['created']}`")
    summary_sect.append(f"- **Số sản phẩm cập nhật thông tin (UPDATE):** `{stats['updated_prices']}`")
    summary_sect.append(f"- **Số sản phẩm giữ nguyên:** `{stats['unchanged']}`")
    summary_sect.append(f"- **Số dòng lỗi:** `{stats['errors']}`\n")
    summary_sect.append("---")
    
    report_lines.insert(3, "\n".join(summary_sect))
    report_lines.extend(sheet_reports)
    
    # Write report file
    report_path = r"D:\Du_An_Bao_Giao\EXCEL_DB_DRYRUN_REPORT.md"
    with open(report_path, "w", encoding="utf-8") as f:
        f.write("\n".join(report_lines))
        
    print("\n" + "="*50)
    print(" SYNCHRONIZATION SUMMARY ")
    print("="*50)
    print(f"Total processed in Excel : {stats['total_excel']}")
    print(f"Created (New INSERT)     : {stats['created']}")
    print(f"Updated (Fields/Prices)  : {stats['updated_prices']}")
    print(f"Unchanged                : {stats['unchanged']}")
    print(f"Errors                   : {stats['errors']}")
    print("="*50)
    print(f"Detailed report saved to: {report_path}")

if __name__ == "__main__":
    if '--real-run' in sys.argv:
        # Run inside a Django transaction block to prevent partial updates in case of errors
        with transaction.atomic():
            main()
    else:
        main()
