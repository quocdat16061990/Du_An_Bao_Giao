"""
Import technical specs from TRA_CUU_TURBO MASTER TỔNG HỢP 2 FILE.xlsx
- Imports SL CHRA, CW, TW, BH as independent products for reference/search
- These products won't have prices, but they have detailed attributes
- Backs up existing images for these if they exist (based on ma_vt)
"""

import sys, os, re, argparse
from pathlib import Path

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'backend.settings')
env_file = Path(__file__).resolve().parent / 'backend' / '.env'
if env_file.exists():
    for line in env_file.read_text(encoding='utf-8').splitlines():
        line = line.strip()
        if line and not line.startswith('#') and '=' in line:
            key, _, val = line.partition('=')
            os.environ.setdefault(key.strip(), val.strip())

import django; django.setup()
from products.models import Product, HangMay, Category, ThuongHieu
from django.utils.text import slugify
from django.db import transaction
import openpyxl

DOCS_DIR = Path(__file__).resolve().parent / 'docs' / 'update-docs-v1'
EXCEL_FILE = DOCS_DIR / 'TRA_CUU_TURBO MASTER TỔNG HỢP 2 FILE.xlsx'

def get_or_create(model_class, ten: str):
    if not ten or ten in ('None', '—', '', ' '): return None
    obj = model_class.objects.filter(ten__iexact=ten).first()
    if obj: return obj
    base_slug = slugify(ten)
    if not base_slug: base_slug = 'khac'

    import time
    try:
        obj, _ = model_class.objects.get_or_create(ten=ten, defaults={'slug': f'{base_slug}-{int(time.time()*1000)}'})
        return obj
    except Exception:
        return None

def clean_val(val):
    if val is None: return ''
    s = str(val).strip()
    if s in ('None', 'NaN', 'nan', '—', '-'): return ''
    return s

def process_master_sheet(ws, start_row, parse_func, stats, image_map):
    for row in range(start_row, ws.max_row + 1):
        data = parse_func(ws, row)
        if not data: continue

        ma_vt = data.get('ma_vt')
        if not ma_vt: continue

        ma_vt = str(ma_vt)[:100]

        if data.get('dry_run'):
            stats['created'] += 1
            continue

        th = get_or_create(ThuongHieu, ten='SL')
        hm = get_or_create(HangMay, ten='Khác')

        try:
            obj, created = Product.objects.update_or_create(
                ma_vt=ma_vt,
                defaults={
                    'loai': data.get('loai', 'so_linh_kien_turbo'),
                    'ten_hang': data.get('ten_hang', '')[:500],
                    'thuong_hieu': th,
                    'hang_may': hm,
                    'is_active': True,
                    'attributes': data.get('attributes', {})
                }
            )
            if created:
                stats['created'] += 1
            else:
                stats['updated'] += 1

            if ma_vt in image_map:
                obj.hinh_anh = image_map[ma_vt]['hinh_anh']
                obj.danh_sach_hinh_anh = image_map[ma_vt]['danh_sach_hinh_anh']
                obj.save(update_fields=['hinh_anh', 'danh_sach_hinh_anh'])
                stats['restored_images'] += 1
        except Exception as e:
            stats['errors'] += 1
            if stats['errors'] <= 3:
                print(f"  ERROR [{ma_vt}]: {e}")

@transaction.atomic
def import_master(dry_run=True):
    if not EXCEL_FILE.exists():
        print(f'[ERROR] File not found: {EXCEL_FILE}')
        return

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    print("--- Backing up images for SL parts ---")
    image_map = {}
    # SL parts usually have MÃ SL starting with 100 or 300, we just backup all just in case
    for p in Product.objects.filter(loai__in=['ruot', 'so_linh_kien_turbo']):
        if p.hinh_anh or p.danh_sach_hinh_anh:
            image_map[p.ma_vt] = {'hinh_anh': p.hinh_anh, 'danh_sach_hinh_anh': p.danh_sach_hinh_anh}

    stats = {'created': 0, 'updated': 0, 'restored_images': 0, 'skipped': 0, 'errors': 0}

    if '⚙️ SL - RUỘT CHRA' in wb.sheetnames:
        print("\n--- Sheet: ⚙️ SL - RUỘT CHRA ---")
        ws = wb['⚙️ SL - RUỘT CHRA']

        def parse_chra(ws, row):
            ma_vt = clean_val(ws.cell(row, 2).value)
            if not ma_vt or not ma_vt.isdigit(): return None
            return {
                'dry_run': dry_run, 'ma_vt': ma_vt, 'loai': 'ruot',
                'ten_hang': f"[SL CHRA] {clean_val(ws.cell(row, 3).value)}",
                'attributes': {
                    'cw_duoi': clean_val(ws.cell(row, 4).value),
                    'cw_dinh': clean_val(ws.cell(row, 5).value),
                    'cw_so_canh': clean_val(ws.cell(row, 6).value),
                    'ten_canh_lua': clean_val(ws.cell(row, 7).value),
                    'tw_duoi': clean_val(ws.cell(row, 8).value),
                    'tw_dinh': clean_val(ws.cell(row, 9).value),
                    'tw_so_canh': clean_val(ws.cell(row, 10).value)
                }
            }
        process_master_sheet(ws, 3, parse_chra, stats, image_map)

    if '🌀 SL - CÁNH GIÓ CW' in wb.sheetnames:
        print("\n--- Sheet: 🌀 SL - CÁNH GIÓ CW ---")
        ws = wb['🌀 SL - CÁNH GIÓ CW']

        def parse_cw(ws, row):
            ma_vt = clean_val(ws.cell(row, 2).value)
            if not ma_vt or not ma_vt.isdigit(): return None
            return {
                'dry_run': dry_run, 'ma_vt': ma_vt, 'loai': 'so_linh_kien_turbo',
                'ten_hang': f"[SL CW] {clean_val(ws.cell(row, 3).value)} - {clean_val(ws.cell(row, 4).value)}",
                'attributes': {
                    'model_turbo': clean_val(ws.cell(row, 4).value),
                    'duoi_a': clean_val(ws.cell(row, 5).value),
                    'dinh_b': clean_val(ws.cell(row, 6).value),
                    'rong_c': clean_val(ws.cell(row, 7).value),
                    'tong_dai_d': clean_val(ws.cell(row, 8).value),
                    'pitch_e': clean_val(ws.cell(row, 9).value),
                    'so_canh_f': clean_val(ws.cell(row, 10).value),
                    'oem': clean_val(ws.cell(row, 12).value)
                }
            }
        process_master_sheet(ws, 3, parse_cw, stats, image_map)

    if '🔥 SL - CÁNH LỬA TW' in wb.sheetnames:
        print("\n--- Sheet: 🔥 SL - CÁNH LỬA TW ---")
        ws = wb['🔥 SL - CÁNH LỬA TW']

        def parse_tw(ws, row):
            ma_vt = clean_val(ws.cell(row, 2).value)
            if not ma_vt or not ma_vt.isdigit(): return None
            return {
                'dry_run': dry_run, 'ma_vt': ma_vt, 'loai': 'so_linh_kien_turbo',
                'ten_hang': f"[SL TW] {clean_val(ws.cell(row, 3).value)} - {clean_val(ws.cell(row, 4).value)}",
                'attributes': {
                    'model_turbo': clean_val(ws.cell(row, 4).value),
                    'duoi_a': clean_val(ws.cell(row, 5).value),
                    'dinh_b': clean_val(ws.cell(row, 6).value),
                    'rong_c': clean_val(ws.cell(row, 7).value),
                    'truc_lon_d': clean_val(ws.cell(row, 8).value),
                    'truc_nho_e': clean_val(ws.cell(row, 9).value),
                    'so_canh_f': clean_val(ws.cell(row, 10).value),
                    'ren_truc_g': clean_val(ws.cell(row, 11).value),
                    'so_piston_ring_h': clean_val(ws.cell(row, 12).value),
                    'chieu_dai_i': clean_val(ws.cell(row, 13).value),
                    'oem': clean_val(ws.cell(row, 14).value)
                }
            }
        process_master_sheet(ws, 3, parse_tw, stats, image_map)

    if '🏠 SL - THÂN BEARING' in wb.sheetnames:
        print("\n--- Sheet: 🏠 SL - THÂN BEARING ---")
        ws = wb['🏠 SL - THÂN BEARING']

        def parse_bh(ws, row):
            ma_vt = clean_val(ws.cell(row, 2).value)
            if not ma_vt or not ma_vt.isdigit(): return None
            return {
                'dry_run': dry_run, 'ma_vt': ma_vt, 'loai': 'so_linh_kien_turbo',
                'ten_hang': f"[SL BH] {clean_val(ws.cell(row, 3).value)} - {clean_val(ws.cell(row, 4).value)}",
                'attributes': {
                    'model_co_ban': clean_val(ws.cell(row, 4).value),
                    'kt_a': clean_val(ws.cell(row, 5).value),
                    'khop_canh_lua_b': clean_val(ws.cell(row, 6).value),
                    'khop_canh_gio_c': clean_val(ws.cell(row, 7).value),
                    'khop_dau_nen_d': clean_val(ws.cell(row, 8).value),
                    'ren_dau_vao_f': clean_val(ws.cell(row, 9).value),
                    'ren_dau_ra_g': clean_val(ws.cell(row, 10).value),
                    'lam_mat': clean_val(ws.cell(row, 11).value)
                }
            }
        process_master_sheet(ws, 3, parse_bh, stats, image_map)

    wb.close()
    print(f"\nImport finished. Stats: {stats}")

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--dry-run', action='store_true', help='Run without modifying DB')
    args = parser.parse_args()

    import_master(dry_run=args.dry_run)
