"""
Import products from BAO_GIA_TURBO_VPS 27-07.xlsx with CORRECT column mapping.
- Backs up existing images by ma_vt
- Deletes ALL existing turbo/ruot/so_linh_kien_turbo products
- Re-imports from Excel
- Restores images to newly created products

Usage: py import_turbo_v1.py --dry-run | py import_turbo_v1.py
"""

import sys, os, re, argparse
from pathlib import Path
from decimal import Decimal, InvalidOperation

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'backend.settings')
env_file = Path(__file__).resolve().parent / 'backend' / '.env'
if env_file.exists():
    for line in env_file.read_text(encoding='utf-8').splitlines():
        line = line.strip()
        if line and not line.startswith('#') and '=' in line:
            key, _, val = line.partition('=')
            os.environ.setdefault(key.strip(), val.strip())

import django; django.setup()
from products.models import Product, HangMay, HangSx, ThuongHieu, Category
from django.utils.text import slugify
import openpyxl

DOCS_DIR = Path(__file__).resolve().parent / 'docs' / 'update-docs-v1'
EXCEL_FILE = DOCS_DIR / 'BAO_GIA_TURBO_VPS 27-07.xlsx'

# Column mapping for BAO_GIA_TURBO (1-indexed)
COL = {
    'hang_may': 1,      # A
    'ma_vt': 2,         # B
    'hang_sx': 3,       # C
    'model_turbo': 4,   # D
    'ma_dong_co': 5,    # E
    'oem_part_no': 6,   # F
    'dac_diem': 7,      # G
    'ung_dung': 8,      # H
    'ghi_chu': 9,       # I
    'thuong_hieu': 10,  # J
    'gia_ban': 11,      # K
    'gia_uu_dai': 12,   # L
    'gia_vip': 13,      # M
    'gia_dl_10': 14,    # N
    'cg_duoi': 15,      # O
    'cg_dinh': 16,      # P
    'cg_so': 17,        # Q
    'cl_duoi': 18,      # R
    'cl_dinh': 19,      # S
    'cl_so': 20,        # T
}

SHEET_LOAI_MAP = {
    '🚗 BÁO GIÁ TURBO ': 'turbo',
    '🔧 RUỘT TURBO (CHRA) ': 'ruot',
    '🔩 SÒ & LINH KIỆN': 'so_linh_kien_turbo',
}

def parse_price(val) -> Decimal | None:
    if val is None: return None
    if isinstance(val, (int, float)): return Decimal(str(int(val)))
    text = str(val).strip()
    if not text or text.lower() in ('none', '—', '-', ''): return None
    clean = re.sub(r'[₫đVND\s,]', '', text)
    if '.' in clean and ',' in clean:
        clean = clean.replace('.', '').replace(',', '.')
    elif '.' in clean:
        parts = clean.split('.')
        if len(parts) > 1 and len(parts[-1]) == 3: clean = clean.replace('.', '')
    elif ',' in clean:
        parts = clean.split(',')
        if len(parts) > 1 and len(parts[-1]) == 3: clean = clean.replace(',', '')
        else: clean = clean.replace(',', '.')
    try: return Decimal(clean)
    except InvalidOperation: return None

def parse_float(val) -> Decimal | None:
    if val is None: return None
    if isinstance(val, (int, float)): return Decimal(str(val))
    try: return Decimal(str(val).strip())
    except InvalidOperation: return None

def get_or_create(model_class, ten: str, **extra):
    obj = model_class.objects.filter(ten__iexact=ten).first()
    if obj: return obj
    slug = slugify(ten)
    return model_class.objects.create(ten=ten, slug=slug, **extra)

def clean_thuong_hieu(raw: str) -> str | None:
    r = raw.strip()
    if not r or r in ('None', '—', '-'): return None
    KNOWN = ['JRONE', 'TBS', 'VIDARIR', 'FIRE', 'EE', 'MX', 'GARRETT', 'SL', 'ISUZU', 'MOBIS', 'SL UK', 'CHÍNH', 'DN-1197']
    r_upper = r.upper()
    for k in KNOWN:
        if k in r_upper: return k
    if re.search(r'\d', r): return None
    return r[:100]

def import_turbo(dry_run=True):
    if not EXCEL_FILE.exists():
        print(f'[ERROR] File not found: {EXCEL_FILE}')
        return

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    types_to_process = list(set(SHEET_LOAI_MAP.values()))

    # Phase 1: Backup images
    print("--- Backing up images ---")
    image_map = {}
    products_to_delete = Product.objects.filter(loai__in=types_to_process)
    for p in products_to_delete:
        if p.hinh_anh or p.danh_sach_hinh_anh:
            image_map[p.ma_vt] = {
                'hinh_anh': p.hinh_anh,
                'danh_sach_hinh_anh': p.danh_sach_hinh_anh
            }
    print(f"Backed up images for {len(image_map)} products.")

    # Phase 2: Delete existing
    print(f'Products to DELETE ({len(types_to_process)} types): {products_to_delete.count()}')
    if not dry_run:
        deleted, _ = products_to_delete.delete()
        print(f'Deleted: {deleted}')

    turbo_cat = get_or_create(Category, ten='Turbo', mo_ta='Bộ turbo tăng áp đầy đủ')
    ruot_cat = get_or_create(Category, ten='Ruột Turbo', mo_ta='Ruột/Core turbo tăng áp')
    slk_cat = get_or_create(Category, ten='Sò & Linh Kiện Turbo', mo_ta='Sò lửa, sò gió & linh kiện turbo')

    loai_cat_map = {
        'turbo': turbo_cat,
        'ruot': ruot_cat,
        'so_linh_kien_turbo': slk_cat,
    }

    stats = {'created': 0, 'restored_images': 0, 'skipped': 0, 'errors': 0}

    for sname, loai in SHEET_LOAI_MAP.items():
        if sname not in wb.sheetnames:
            print(f'[WARN] Sheet not found: {sname}')
            continue

        ws = wb[sname]
        cat = loai_cat_map[loai]
        print(f'\n--- {sname} -> loai={loai} ---')

        for row in range(5, ws.max_row + 1):
            ma_vt = str(ws.cell(row, COL['ma_vt']).value or '').strip()
            if not ma_vt.startswith('HH'):
                continue

            hang_may_raw = str(ws.cell(row, COL['hang_may']).value or '').strip().replace('└─', '').strip()
            if not hang_may_raw or hang_may_raw in ('None', '—', ''):
                stats['skipped'] += 1
                continue

            hang_sx_raw = str(ws.cell(row, COL['hang_sx']).value or '').strip()
            thuong_hieu_raw = str(ws.cell(row, COL['thuong_hieu']).value or '').strip()
            model_turbo = str(ws.cell(row, COL['model_turbo']).value or '').strip()
            ma_dong_co = str(ws.cell(row, COL['ma_dong_co']).value or '').strip()
            oem_part_no = str(ws.cell(row, COL['oem_part_no']).value or '').strip()
            dac_diem = str(ws.cell(row, COL['dac_diem']).value or '').strip()
            ung_dung = str(ws.cell(row, COL['ung_dung']).value or '').strip()
            ghi_chu = str(ws.cell(row, COL['ghi_chu']).value or '').strip()

            gia_ban = parse_price(ws.cell(row, COL['gia_ban']).value)
            gia_uu_dai = parse_price(ws.cell(row, COL['gia_uu_dai']).value)
            gia_vip = parse_price(ws.cell(row, COL['gia_vip']).value)
            gia_dl_10 = parse_price(ws.cell(row, COL['gia_dl_10']).value)

            cg_duoi = parse_float(ws.cell(row, COL['cg_duoi']).value)
            cg_dinh = parse_float(ws.cell(row, COL['cg_dinh']).value)
            cg_so = str(ws.cell(row, COL['cg_so']).value or '').strip()
            cl_duoi = parse_float(ws.cell(row, COL['cl_duoi']).value)
            cl_dinh = parse_float(ws.cell(row, COL['cl_dinh']).value)
            cl_so = str(ws.cell(row, COL['cl_so']).value or '').strip()

            if dry_run:
                stats['created'] += 1
                continue

            hm = get_or_create(HangMay, ten=hang_may_raw)
            hsx = get_or_create(HangSx, ten=hang_sx_raw) if hang_sx_raw and hang_sx_raw not in ('None', '—', '-') else None
            th_clean = clean_thuong_hieu(thuong_hieu_raw)
            th = get_or_create(ThuongHieu, ten=th_clean) if th_clean else None

            try:
                obj = Product.objects.create(
                    ma_vt=ma_vt,
                    loai=loai,
                    category=cat,
                    hang_may=hm,
                    hang_sx=hsx,
                    thuong_hieu=th,
                    model_turbo=model_turbo[:300],
                    ma_dong_co=ma_dong_co[:300],
                    oem_part_no=oem_part_no,
                    dac_diem=dac_diem,
                    ung_dung=ung_dung,
                    ghi_chu=ghi_chu,
                    gia_dai_ly=gia_ban,
                    gia_uu_dai=gia_uu_dai,
                    gia_vip=gia_vip,
                    gia_dl_10=gia_dl_10,
                    cg_duoi=cg_duoi,
                    cg_dinh=cg_dinh,
                    cg_so=cg_so[:20],
                    cl_duoi=cl_duoi,
                    cl_dinh=cl_dinh,
                    cl_so=cl_so[:20],
                    is_active=True,
                )
                stats['created'] += 1

                # Restore images
                if ma_vt in image_map:
                    obj.hinh_anh = image_map[ma_vt]['hinh_anh']
                    obj.danh_sach_hinh_anh = image_map[ma_vt]['danh_sach_hinh_anh']
                    obj.save(update_fields=['hinh_anh', 'danh_sach_hinh_anh'])
                    stats['restored_images'] += 1
            except Exception as e:
                stats['errors'] += 1
                if stats['errors'] <= 3:
                    print(f'  ERROR [{ma_vt}]: {e}')

    wb.close()
    print(f"\nImport finished. Stats: {stats}")

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--dry-run', action='store_true', help='Run without modifying DB')
    args = parser.parse_args()

    import_turbo(dry_run=args.dry_run)
